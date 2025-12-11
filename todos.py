#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para generar análisis cruzado de TODAS las preguntas desde P3 en adelante
con todas las variables demográficas y de clasificación.

Autor: Generado automáticamente
Fecha: 2025
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
import re

def crear_rango_edad(edad):
    """
    Crea rangos de edad a partir de la edad numérica.
    """
    if pd.isna(edad):
        return None
    try:
        edad_num = int(float(edad))
        if edad_num <= 25:
            return '18 - 25'
        elif edad_num <= 35:
            return '26 - 35'
        elif edad_num <= 45:
            return '36 - 45'
        elif edad_num <= 60:
            return '46 - 60'
        else:
            return 'Más de 61'
    except:
        return None

def obtener_region_oficina(oficina):
    """
    Agrupa las oficinas/agencias/delegaciones por región.
    """
    if pd.isna(oficina):
        return None
    
    oficina_str = str(oficina).strip()
    
    # Central
    central = ['Chimaltenango', 'El Progreso', 'Guatemala', 'Sacatepéquez']
    if oficina_str in central:
        return 'Central'
    
    # Occidente
    occidente = ['Huehuetenango', 'Quetzaltenango', 'Quiché', 'San Marcos', 'Sololá', 'Totonicapán']
    if oficina_str in occidente:
        return 'Occidente'
    
    # Sur
    sur = ['Escuintla', 'Jutiapa', 'Retalhuleu', 'Santa Rosa', 'Suchitepéquez']
    if oficina_str in sur:
        return 'Sur'
    
    # Nororiente
    nororiente = ['Alta Verapaz', 'Baja Verapaz', 'Chiquimula', 'Izabal', 'Jalapa', 'Petén', 'Zacapa']
    if oficina_str in nororiente:
        return 'Nororiente'
    
    return None

def obtener_region_aduana(aduana):
    """
    Agrupa las aduanas por región.
    """
    if pd.isna(aduana):
        return None
    
    aduana_str = str(aduana).strip()
    
    # Central
    if 'Central Guatemala' in aduana_str:
        return 'Central'
    
    # Occidente
    if 'El Carmen' in aduana_str or 'La Mesilla' in aduana_str:
        return 'Occidente'
    
    # Sur
    if 'San Cristóbal' in aduana_str or 'Valle Nuevo' in aduana_str or 'Puerto Quetzal' in aduana_str:
        return 'Sur'
    
    # Nororiente
    if 'Integrada Corinto' in aduana_str or 'Integrada El Florido' in aduana_str or 'Puerto Barrios' in aduana_str or 'Santo Tomás' in aduana_str or 'Tikal' in aduana_str:
        return 'Nororiente'
    
    return None

def detectar_combinaciones_multiples(df, columna):
    """
    Detecta si una columna tiene combinaciones múltiples (valores con comas).
    """
    if columna not in df.columns:
        return False
    
    # EXCEPCIÓN: P43 tiene comas pero es una sola opción, no múltiple
    if columna == 'P43 - Tipo de Punto':
        return False
    
    valores = df[columna].dropna()
    if len(valores) == 0:
        return False
    
    valores_con_comas = valores.astype(str).str.contains(',', na=False)
    return valores_con_comas.sum() > 0

def obtener_opciones_unicas(df, columna, tiene_combinaciones):
    """
    Obtiene las opciones únicas de una pregunta.
    Si tiene combinaciones, extrae las opciones principales (que aparecen más frecuentemente).
    Si no tiene combinaciones, devuelve los valores únicos.
    """
    if columna not in df.columns:
        return []
    
    valores = df[columna].dropna()
    if len(valores) == 0:
        return []
    
    if tiene_combinaciones:
        # Extraer todas las opciones individuales de las combinaciones
        opciones_contador = {}
        for valor in valores:
            valor_str = str(valor).strip()
            # Dividir por comas y limpiar
            partes = [p.strip() for p in valor_str.split(',')]
            for parte in partes:
                if parte:
                    opciones_contador[parte] = opciones_contador.get(parte, 0) + 1
        
        # Ordenar por frecuencia y devolver las más comunes
        # Para preguntas con formato "a. Opción", "b. Opción", etc., estas serán las principales
        opciones_ordenadas = sorted(opciones_contador.items(), key=lambda x: x[1], reverse=True)
        
        # Si hay opciones que empiezan con letra y punto (a., b., c., etc.), priorizarlas
        opciones_principales = [op for op, count in opciones_ordenadas if re.match(r'^[a-z]\.\s', op)]
        opciones_secundarias = [op for op, count in opciones_ordenadas if not re.match(r'^[a-z]\.\s', op)]
        
        # Combinar: primero las principales (ordenadas por letra), luego las secundarias
        if opciones_principales:
            # Ordenar las principales por letra
            opciones_principales.sort(key=lambda x: x[0] if x else 'z')
            return opciones_principales + [op for op, _ in opciones_ordenadas if op not in opciones_principales]
        else:
            # Si no hay formato a., b., c., devolver todas ordenadas por frecuencia
            return [op for op, _ in opciones_ordenadas]
    else:
        # Sin combinaciones, devolver valores únicos
        return sorted(valores.unique().tolist())

def normalizar_combinaciones(valor, opciones_principales):
    """
    Normaliza valores con combinaciones múltiples.
    Detecta qué opciones principales están presentes y las ordena.
    """
    if pd.isna(valor):
        return None
    valor_str = str(valor).strip()
    
    opciones_presentes = []
    for opcion in opciones_principales:
        if opcion in valor_str:
            opciones_presentes.append(opcion)
    
    if len(opciones_presentes) == 0:
        return None
    
    return ', '.join(opciones_presentes)

def redondear_porcentaje(valor):
    """Redondea porcentajes: si tiene .5 o más, redondea hacia arriba"""
    if pd.isna(valor) or valor == 0:
        return 0
    # Si el decimal es >= 0.5, redondear hacia arriba
    if valor % 1 >= 0.5:
        return int(valor) + 1
    else:
        return int(valor)

def generar_hoja_pregunta(wb, df, pregunta_num, pregunta_col, pregunta_nombre, tiene_combinaciones):
    """
    Genera una hoja completa para una pregunta específica.
    """
    print(f"\n{'='*80}")
    print(f"Procesando {pregunta_nombre}")
    print(f"{'='*80}")
    
    # Crear hoja
    ws = wb.create_sheet(title=f"P{pregunta_num}")
    
    # Obtener opciones de la pregunta
    opciones = obtener_opciones_unicas(df, pregunta_col, tiene_combinaciones)
    
    if len(opciones) == 0:
        print(f"  ⚠ No se encontraron opciones para {pregunta_nombre}")
        return
    
    print(f"  Opciones encontradas: {len(opciones)}")
    if tiene_combinaciones:
        print(f"  Tipo: Con combinaciones múltiples")
    else:
        print(f"  Tipo: Sin combinaciones múltiples")
    
    # Preparar datos
    df_work = df.copy()
    
    # FILTRO ESPECIAL PARA P6: Solo incluir registros con "b. Contact Center" en P3
    if pregunta_col == 'P6 - Gestión Contact Center':
        p3_col = 'P3 - Medios SAT Utilizados'
        if p3_col in df_work.columns:
            # Filtrar solo registros que tienen "b. Contact Center" en P3
            mask_p3_contact = df_work[p3_col].astype(str).str.contains('b. Contact Center', na=False)
            df_work = df_work[mask_p3_contact].copy()
            print(f"  ⚠ P6 es condicional: Filtrando solo registros con 'b. Contact Center' en P3")
            print(f"  Registros después del filtro: {len(df_work)}")
    
    # FILTRO ESPECIAL PARA P7: Solo incluir registros con "b. Contact Center" en P3
    if pregunta_col == 'P7 - Medio Contact Center':
        p3_col = 'P3 - Medios SAT Utilizados'
        if p3_col in df_work.columns:
            # Filtrar solo registros que tienen "b. Contact Center" en P3
            mask_p3_contact = df_work[p3_col].astype(str).str.contains('b. Contact Center', na=False)
            df_work = df_work[mask_p3_contact].copy()
            print(f"  ⚠ P7 es condicional: Filtrando solo registros con 'b. Contact Center' en P3")
            print(f"  Registros después del filtro: {len(df_work)}")
    
    # FILTRO ESPECIAL PARA P8: Solo incluir registros con "a. Presencial" en P3
    if pregunta_col == 'P8 - Gestión Visita Presencial':
        p3_col = 'P3 - Medios SAT Utilizados'
        if p3_col in df_work.columns:
            # Filtrar solo registros que tienen "a. Presencial" en P3
            mask_p3_presencial = df_work[p3_col].astype(str).str.contains('a. Presencial', na=False)
            df_work = df_work[mask_p3_presencial].copy()
            print(f"  ⚠ P8 es condicional: Filtrando solo registros con 'a. Presencial' en P3")
            print(f"  Registros después del filtro: {len(df_work)}")
    
    # Crear rangos de edad
    df_work['Rango_Edad'] = df_work['P36 - Edad'].apply(crear_rango_edad)
    
    # Crear regiones
    df_work['Region_Oficina'] = df_work['P44 - Oficina/Agencia/Delegación'].apply(obtener_region_oficina)
    df_work['Region_Aduana'] = df_work['P44 - Aduana'].apply(obtener_region_aduana)
    
    # Si tiene combinaciones, normalizar
    if tiene_combinaciones:
        df_work[f'{pregunta_col}_norm'] = df_work[pregunta_col].apply(
            lambda x: normalizar_combinaciones(x, opciones)
        )
    
    # Definir estilos
    thin_side = Side(style='thin', color='FFD0D0D0')
    medium_side = Side(style='medium')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    fill_fila1 = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    fill_header = PatternFill(start_color='FFE7E6E6', end_color='FFE7E6E6', fill_type='solid')
    
    # Fila 1: Título de la pregunta
    ws.cell(row=1, column=1, value=pregunta_nombre)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=1, column=1).fill = fill_fila1
    
    # Definir variables de cruce (igual que P3/P4)
    col = 1
    variables = {
        'P37 Género': {
            'columna': 'P37 - Género',
            'categorias': ['H', 'M', 'No deseo responder'],
            'col_inicio': col + 2
        },
        'P3 Medios SAT utilizados': {
            'columna': 'P3 - Medios SAT Utilizados',
            'categorias': ['a. Presencial', 'b. Contact Center', 'c. Servicios Electrónicos'],
            'col_inicio': col + 5,
            'usa_contains': True  # Marcar que usa str.contains() para contar
        },
        'Rango de edad': {
            'columna': 'Rango_Edad',
            'categorias': ['18 - 25', '26 - 35', '36 - 45', '46 - 60', 'Más de 61'],
            'col_inicio': col + 8
        },
        'P40 Nivel académico': {
            'columna': 'P40 - Nivel Académico',
            'categorias': [
                'a. Ninguno', 'b. Primaria incompleta', 'c. Primaria completa',
                'd. Secundaria incompleta (1ro a 3ro básico)', 'e. Secundaria Completa (1ro a 3ro básico)',
                'f. Diversificado incompleto', 'g. Diversificado completo', 'h. Técnico',
                'i. Universidad incompleta', 'j. Universidad Completa', 'k. Maestría / Posgrado'
            ],
            'col_inicio': col + 13
        },
        'P39 Idiomas': {
            'columna': 'P39 - Idiomas',
            'categorias': [
                'a. Achi', 'b. Qánjob\'al', 'c. Q\'eqchi', 'd. Akateco', 'e. Kaqchikel',
                'f. Sakapulteko', 'h. Kiché', 'i. Sipakapense', 'k. Mam', 'n. Mopan',
                'p. Ixil', 'q. Poqomam', 's. Jakalteco', 't. Poqomchi', 'u. Ninguno',
                'v. Inglés', 'w. Otro'
            ],
            'col_inicio': col + 24
        },
        'P44 Oficina/Agencia/Delegación': {
            'columna': 'P44 - Oficina/Agencia/Delegación',
            'categorias': [
                'Alta Verapaz', 'Baja Verapaz', 'Chimaltenango', 'Chiquimula', 'El Progreso',
                'Escuintla', 'Guatemala', 'Huehuetenango', 'Izabal', 'Jalapa', 'Jutiapa',
                'Petén', 'Quetzaltenango', 'Quiché', 'Retalhuleu', 'Sacatepéquez', 'San Marcos',
                'Santa Rosa', 'Sololá', 'Suchitepéquez', 'Totonicapán', 'Zacapa'
            ],
            'col_inicio': col + 41
        },
        'P44.1 Aduana': {
            'columna': 'P44 - Aduana',
            'categorias': [
                'Central Guatemala', 'El Carmen', 'Integrada Corinto', 'Integrada El Florido',
                'La Mesilla', 'Puerto Barrios Almacenadora Pelícano, S.A -ALPELSA', 'Puerto Quetzal',
                'San Cristóbal', 'Santo Tomás de Castilla Zona Libre de Industria y Comercio -ZOLIC-',
                'Tikal', 'Valle Nuevo'
            ],
            'col_inicio': col + 63
        },
        'P9 Personería': {
            'columna': 'P9 - Personería',
            'categorias': [
                'a. Contribuyente/Propietario.', 'b. Representante Legal', 'c. Abogado y Notario',
                'd. Mandatario', 'e. Contador/auxiliar', 'f. Contador Público y Auditor',
                'g. Gestor Tributario', 'h. Importador', 'i. Exportador', 'j. Asistente de Agente',
                'k. Auxiliar Gestor Tributario', 'm. Consolidador/Descons.', 'n. Transportista Ad',
                'p. Mensajero', 'r. Otro'
            ],
            'col_inicio': col + 74
        },
        'P38 Etnia': {
            'columna': 'P38 - Etnia',
            'categorias': ['Garifuna', 'Ladino', 'Maya', 'Otro', 'Xinca'],
            'col_inicio': col + 89
        },
        'Oficina/Agencia/Delegación': {
            'columna': 'Region_Oficina',
            'categorias': ['Central', 'Occidente', 'Sur', 'Nororiente'],
            'col_inicio': col + 94
        },
        'Aduana': {
            'columna': 'Region_Aduana',
            'categorias': ['Central', 'Occidente', 'Sur', 'Nororiente'],
            'col_inicio': col + 98
        }
    }
    
    # Calcular total de columnas
    total_columnas = 2  # Columna vacía + TOTAL
    for var_nombre, var_info in variables.items():
        total_columnas += len(var_info['categorias'])
    
    # Aplicar merge y bordes al título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columnas)
    ws.cell(row=1, column=1).border = Border(
        left=medium_side,
        right=medium_side,
        top=medium_side,
        bottom=thin_side
    )
    
    # Fila 2: Vacía con fondo gris
    for col_idx in range(1, total_columnas + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = fill_fila1
        cell.border = Border(
            left=thin_side, 
            right=thin_side, 
            top=thin_side, 
            bottom=thin_side
        )
    
    # Fila 3: Encabezados principales
    col_actual = 1
    ws.cell(row=3, column=col_actual, value='')
    ws.cell(row=3, column=col_actual).fill = fill_header
    ws.cell(row=3, column=col_actual).border = Border(left=medium_side, right=medium_side, top=thin_side, bottom=thin_side)
    col_actual += 1
    
    ws.cell(row=3, column=col_actual, value='TOTAL')
    ws.cell(row=3, column=col_actual).fill = fill_header
    ws.cell(row=3, column=col_actual).border = Border(left=medium_side, right=medium_side, top=thin_side, bottom=thin_side)
    ws.cell(row=3, column=col_actual).font = Font(bold=True)
    ws.cell(row=3, column=col_actual).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    col_actual += 1
    
    for var_nombre, var_info in variables.items():
        num_cols = len(var_info['categorias'])
        inicio = col_actual
        fin = col_actual + num_cols - 1
        
        ws.merge_cells(start_row=3, start_column=inicio, end_row=3, end_column=fin)
        ws.cell(row=3, column=inicio, value=var_nombre)
        ws.cell(row=3, column=inicio).fill = fill_header
        ws.cell(row=3, column=inicio).border = Border(left=medium_side, right=medium_side, top=thin_side, bottom=thin_side)
        ws.cell(row=3, column=inicio).font = Font(bold=True)
        ws.cell(row=3, column=inicio).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        col_actual = fin + 1
    
    # Fila 4: Sub-encabezados
    col_actual = 1
    ws.cell(row=4, column=col_actual, value='')
    ws.cell(row=4, column=col_actual).border = thin_border
    col_actual += 1
    
    ws.cell(row=4, column=col_actual, value='')
    ws.cell(row=4, column=col_actual).border = Border(
        left=medium_side, 
        right=medium_side, 
        top=Side(style='medium'), 
        bottom=Side(style='thin', color='FFD0D0D0')
    )
    col_actual += 1
    
    for var_nombre, var_info in variables.items():
        num_cats = len(var_info['categorias'])
        for i, cat in enumerate(var_info['categorias']):
            es_primera = (i == 0)
            es_ultima = (i == num_cats - 1)
            ws.cell(row=4, column=col_actual, value=cat)
            ws.cell(row=4, column=col_actual).font = Font(bold=True)
            border = Border(
                left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),
                top=Side(style='medium'),
                bottom=Side(style='thin', color='FFD0D0D0')
            )
            ws.cell(row=4, column=col_actual).border = border
            ws.cell(row=4, column=col_actual).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            col_actual += 1
    
    # Filas de datos
    fila = 5
    
    # Filtrar solo registros con respuesta a esta pregunta
    if tiene_combinaciones:
        # Para combinaciones, incluir todos los registros que tengan al menos una opción
        # Usar la columna original, no la normalizada
        df_pregunta = df_work[df_work[pregunta_col].notna()].copy()
    else:
        df_pregunta = df_work[df_work[pregunta_col].notna()].copy()
    
    for idx_opcion, opcion in enumerate(opciones):
        col_actual = 1
        
        # Nombre de la fila
        ws.cell(row=fila, column=col_actual, value=opcion)
        border = Border(
            left=Side(style='thin', color='FFD0D0D0'),
            right=Side(style='thin', color='FFD0D0D0'),
            top=Side(style='medium' if idx_opcion == 0 else 'thin', color='FFD0D0D0'),
            bottom=Side(style='thin', color='FFD0D0D0')
        )
        ws.cell(row=fila, column=col_actual).border = border
        ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='left', vertical='center')
        col_actual += 1
        
        # TOTAL
        if tiene_combinaciones:
            total = len(df_work[df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)])
        else:
            total = len(df_work[df_work[pregunta_col] == opcion])
        
        ws.cell(row=fila, column=col_actual, value=total)
        border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium' if idx_opcion == 0 else 'thin', color='FFD0D0D0'),
            bottom=Side(style='thin', color='FFD0D0D0')
        )
        ws.cell(row=fila, column=col_actual).border = border
        ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
        col_actual += 1
        
        # Datos por variable
        for var_nombre, var_info in variables.items():
            col_original = var_info['columna']
            num_cats = len(var_info['categorias'])
            
            for i, cat in enumerate(var_info['categorias']):
                es_primera = (i == 0)
                es_ultima = (i == num_cats - 1)
                
                # Contar intersección
                if tiene_combinaciones:
                    if col_original == 'Rango_Edad':
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original] == cat)
                        ])
                    elif col_original == 'P39 - Idiomas':
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original].astype(str).str.contains(cat, na=False))
                        ])
                    elif col_original == 'P3 - Medios SAT Utilizados':
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original].astype(str).str.contains(cat, na=False))
                        ])
                    elif col_original == 'Region_Oficina' or col_original == 'Region_Aduana':
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original] == cat)
                        ])
                    else:
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original] == cat)
                        ])
                else:
                    if col_original == 'Rango_Edad':
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original] == cat)
                        ])
                    elif col_original == 'P39 - Idiomas':
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original].astype(str).str.contains(cat, na=False))
                        ])
                    elif col_original == 'P3 - Medios SAT Utilizados':
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original].astype(str).str.contains(cat, na=False))
                        ])
                    elif col_original == 'Region_Oficina' or col_original == 'Region_Aduana':
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original] == cat)
                        ])
                    else:
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original] == cat)
                        ])
                
                ws.cell(row=fila, column=col_actual, value=count)
                border = Border(
                    left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                    right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),
                    top=Side(style='medium' if idx_opcion == 0 else 'thin', color='FFD0D0D0'),
                    bottom=Side(style='thin', color='FFD0D0D0')
                )
                ws.cell(row=fila, column=col_actual).border = border
                ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
                col_actual += 1
        
        fila += 1
    
    # Asegurar bordes right=medium en última columna
    for row in range(1, fila + 1):
        cell = ws.cell(row=row, column=total_columnas)
        if cell.border:
            current_border = cell.border
            new_border = Border(
                left=current_border.left,
                right=Side(style='medium'),
                top=current_border.top,
                bottom=current_border.bottom
            )
            cell.border = new_border
    
    # Fila TOTAL
    col_actual = 1
    ws.cell(row=fila, column=col_actual, value='TOTAL')
    ws.cell(row=fila, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila, column=col_actual).border = Border(
        left=Side(style='thin', color='FFD0D0D0'),
        right=Side(style='thin', color='FFD0D0D0'),
        top=Side(style='thin', color='FFD0D0D0'),
        bottom=Side(style='medium')
    )
    ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
    col_actual += 1
    
    # TOTAL general
    total_general = len(df_pregunta)
    ws.cell(row=fila, column=col_actual, value=total_general)
    ws.cell(row=fila, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila, column=col_actual).border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='thin', color='FFD0D0D0'),
        bottom=Side(style='medium')
    )
    ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
    col_actual += 1
    
    # Totales por categoría
    for var_nombre, var_info in variables.items():
        col_original = var_info['columna']
        num_cats = len(var_info['categorias'])
        
        for i, cat in enumerate(var_info['categorias']):
            es_primera = (i == 0)
            es_ultima = (i == num_cats - 1)
            
            if col_original == 'Rango_Edad':
                total_cat = len(df_pregunta[df_pregunta[col_original] == cat])
            elif col_original == 'P39 - Idiomas':
                total_cat = len(df_pregunta[df_pregunta[col_original].astype(str).str.contains(cat, na=False)])
            elif col_original == 'P3 - Medios SAT Utilizados':
                total_cat = len(df_pregunta[df_pregunta[col_original].astype(str).str.contains(cat, na=False)])
            elif col_original == 'Region_Oficina' or col_original == 'Region_Aduana':
                total_cat = len(df_pregunta[df_pregunta[col_original] == cat])
            else:
                total_cat = len(df_pregunta[df_pregunta[col_original] == cat])
            
            ws.cell(row=fila, column=col_actual, value=total_cat)
            ws.cell(row=fila, column=col_actual).font = Font(bold=True)
            border = Border(
                left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),
                top=Side(style='thin', color='FFD0D0D0'),
                bottom=Side(style='medium')
            )
            ws.cell(row=fila, column=col_actual).border = border
            ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
            col_actual += 1
    
    # Asegurar bordes en última fila
    for col_idx in range(1, total_columnas + 1):
        cell = ws.cell(row=fila, column=col_idx)
        if cell.border:
            current_border = cell.border
            right_style = Side(style='medium') if col_idx == total_columnas else current_border.right
            new_border = Border(
                left=current_border.left,
                right=right_style,
                top=current_border.top,
                bottom=Side(style='medium')
            )
            cell.border = new_border
    
    # Agregar dos filas vacías
    fila += 2
    
    # TABLA DE PORCENTAJES
    fila_porcentajes = fila
    
    # Fila de encabezados principales
    col_actual = 1
    ws.cell(row=fila_porcentajes, column=col_actual, value='')
    ws.cell(row=fila_porcentajes, column=col_actual).fill = fill_header
    ws.cell(row=fila_porcentajes, column=col_actual).border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=thin_side)
    col_actual += 1
    
    ws.cell(row=fila_porcentajes, column=col_actual, value='TOTAL')
    ws.cell(row=fila_porcentajes, column=col_actual).fill = fill_header
    ws.cell(row=fila_porcentajes, column=col_actual).border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=thin_side)
    ws.cell(row=fila_porcentajes, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    col_actual += 1
    
    for var_nombre, var_info in variables.items():
        num_cols = len(var_info['categorias'])
        inicio = col_actual
        fin = col_actual + num_cols - 1
        
        ws.merge_cells(start_row=fila_porcentajes, start_column=inicio, end_row=fila_porcentajes, end_column=fin)
        ws.cell(row=fila_porcentajes, column=inicio, value=var_nombre)
        ws.cell(row=fila_porcentajes, column=inicio).fill = fill_header
        ws.cell(row=fila_porcentajes, column=inicio).border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=thin_side)
        ws.cell(row=fila_porcentajes, column=inicio).font = Font(bold=True)
        ws.cell(row=fila_porcentajes, column=inicio).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        col_actual = fin + 1
    
    # Fila de sub-encabezados
    fila_porcentajes += 1
    col_actual = 1
    ws.cell(row=fila_porcentajes, column=col_actual, value='')
    ws.cell(row=fila_porcentajes, column=col_actual).border = thin_border
    col_actual += 1
    
    ws.cell(row=fila_porcentajes, column=col_actual, value='')
    ws.cell(row=fila_porcentajes, column=col_actual).border = Border(
        left=medium_side, 
        right=medium_side, 
        top=Side(style='medium'), 
        bottom=Side(style='thin', color='FFD0D0D0')
    )
    col_actual += 1
    
    for var_nombre, var_info in variables.items():
        num_cats = len(var_info['categorias'])
        for i, cat in enumerate(var_info['categorias']):
            es_primera = (i == 0)
            es_ultima = (i == num_cats - 1)
            ws.cell(row=fila_porcentajes, column=col_actual, value=cat)
            ws.cell(row=fila_porcentajes, column=col_actual).font = Font(bold=True)
            border = Border(
                left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),
                top=Side(style='medium'),
                bottom=Side(style='thin', color='FFD0D0D0')
            )
            ws.cell(row=fila_porcentajes, column=col_actual).border = border
            ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            col_actual += 1
    
    # Filas de datos con porcentajes
    fila_porcentajes += 1
    for idx_opcion, opcion in enumerate(opciones):
        col_actual = 1
        
        # Nombre de la fila
        ws.cell(row=fila_porcentajes, column=col_actual, value=opcion)
        border = Border(
            left=Side(style='thin', color='FFD0D0D0'),
            right=Side(style='thin', color='FFD0D0D0'),
            top=Side(style='medium' if idx_opcion == 0 else 'thin', color='FFD0D0D0'),
            bottom=Side(style='thin', color='FFD0D0D0')
        )
        ws.cell(row=fila_porcentajes, column=col_actual).border = border
        ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='left', vertical='center')
        col_actual += 1
        
        # TOTAL - calcular porcentaje
        if tiene_combinaciones:
            total_absoluto = len(df_work[df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)])
        else:
            total_absoluto = len(df_work[df_work[pregunta_col] == opcion])
        
        porcentaje_total = (total_absoluto / total_general * 100) if total_general > 0 else 0
        # Truncar a dos decimales sin redondear (almacenar como decimal)
        # Ejemplo: 13.456% -> int(13.456 * 100) = 1345 -> 1345/10000 = 0.1345 (representa 13.45%)
        porcentaje_decimal = int(porcentaje_total * 100) / 10000
        
        if porcentaje_decimal == 0:
            ws.cell(row=fila_porcentajes, column=col_actual, value="---")
        else:
            ws.cell(row=fila_porcentajes, column=col_actual, value=porcentaje_decimal)
            ws.cell(row=fila_porcentajes, column=col_actual).number_format = '0.00%'
        
        border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium' if idx_opcion == 0 else 'thin', color='FFD0D0D0'),
            bottom=Side(style='thin', color='FFD0D0D0')
        )
        ws.cell(row=fila_porcentajes, column=col_actual).border = border
        ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
        col_actual += 1
        
        # Datos por variable (porcentajes)
        for var_nombre, var_info in variables.items():
            col_original = var_info['columna']
            num_cats = len(var_info['categorias'])
            
            for i, cat in enumerate(var_info['categorias']):
                es_primera = (i == 0)
                es_ultima = (i == num_cats - 1)
                
                # Contar intersección
                if tiene_combinaciones:
                    if col_original == 'Rango_Edad':
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original] == cat)
                        ])
                    elif col_original == 'P39 - Idiomas':
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original].astype(str).str.contains(cat, na=False))
                        ])
                    elif col_original == 'P3 - Medios SAT Utilizados':
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original].astype(str).str.contains(cat, na=False))
                        ])
                    elif col_original == 'Region_Oficina' or col_original == 'Region_Aduana':
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original] == cat)
                        ])
                    else:
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original] == cat)
                        ])
                else:
                    if col_original == 'Rango_Edad':
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original] == cat)
                        ])
                    elif col_original == 'P39 - Idiomas':
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original].astype(str).str.contains(cat, na=False))
                        ])
                    elif col_original == 'P3 - Medios SAT Utilizados':
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original].astype(str).str.contains(cat, na=False))
                        ])
                    elif col_original == 'Region_Oficina' or col_original == 'Region_Aduana':
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original] == cat)
                        ])
                    else:
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original] == cat)
                        ])
                
                # Calcular porcentaje VERTICAL (sobre el total de esa categoría)
                if col_original == 'Rango_Edad':
                    total_categoria = len(df_pregunta[df_pregunta[col_original] == cat])
                elif col_original == 'P39 - Idiomas':
                    total_categoria = len(df_pregunta[df_pregunta[col_original].astype(str).str.contains(cat, na=False)])
                elif col_original == 'P3 - Medios SAT Utilizados':
                    total_categoria = len(df_pregunta[df_pregunta[col_original].astype(str).str.contains(cat, na=False)])
                elif col_original == 'Region_Oficina' or col_original == 'Region_Aduana':
                    total_categoria = len(df_pregunta[df_pregunta[col_original] == cat])
                else:
                    total_categoria = len(df_pregunta[df_pregunta[col_original] == cat])
                
                porcentaje = (count / total_categoria * 100) if total_categoria > 0 else 0
                # Truncar a dos decimales sin redondear (almacenar como decimal con 2 decimales)
                porcentaje_decimal = int(porcentaje * 100) / 10000
                
                if porcentaje_decimal == 0:
                    ws.cell(row=fila_porcentajes, column=col_actual, value="---")
                else:
                    ws.cell(row=fila_porcentajes, column=col_actual, value=porcentaje_decimal)
                    ws.cell(row=fila_porcentajes, column=col_actual).number_format = '0.00%'
                
                border = Border(
                    left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                    right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),
                    top=Side(style='medium' if idx_opcion == 0 else 'thin', color='FFD0D0D0'),
                    bottom=Side(style='thin', color='FFD0D0D0')
                )
                ws.cell(row=fila_porcentajes, column=col_actual).border = border
                ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
                col_actual += 1
        
        fila_porcentajes += 1
    
    # Fila TOTAL de porcentajes
    col_actual = 1
    ws.cell(row=fila_porcentajes, column=col_actual, value='TOTAL')
    ws.cell(row=fila_porcentajes, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila_porcentajes, column=col_actual).border = Border(
        left=Side(style='thin', color='FFD0D0D0'),
        right=Side(style='thin', color='FFD0D0D0'),
        top=Side(style='thin', color='FFD0D0D0'),
        bottom=Side(style='medium')
    )
    ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
    col_actual += 1
    
    # TOTAL general - suma VERTICAL de porcentajes (suma de los porcentajes de arriba en esta columna)
    suma_total = 0
    for idx_opcion in range(len(opciones)):
        fila_anterior = fila_porcentajes - len(opciones) + idx_opcion
        valor_celda = ws.cell(row=fila_anterior, column=col_actual).value
        if valor_celda is not None and valor_celda != "---":
            if isinstance(valor_celda, (int, float)) and valor_celda > 0:
                # Los valores están en formato decimal (0.13), multiplicar por 10000 para obtener centésimas
                suma_total += int(valor_celda * 10000)
            elif isinstance(valor_celda, str) and valor_celda != "---" and '%' in valor_celda:
                # Extraer el número del porcentaje
                try:
                    num_porcentaje = float(valor_celda.replace('%', '').strip())
                    suma_total += int(num_porcentaje * 100)  # Convertir a centésimas
                except:
                    pass
    
    if suma_total == 0:
        ws.cell(row=fila_porcentajes, column=col_actual, value="---")
    else:
        # Truncar a dos decimales sin redondear (almacenar como decimal con 2 decimales)
        total_decimal = int(suma_total) / 10000
        ws.cell(row=fila_porcentajes, column=col_actual, value=total_decimal)
        ws.cell(row=fila_porcentajes, column=col_actual).number_format = '0.00%'
    
    ws.cell(row=fila_porcentajes, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila_porcentajes, column=col_actual).border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='thin', color='FFD0D0D0'),
        bottom=Side(style='medium')
    )
    ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
    col_actual += 1
    
    # Totales por categoría - suma vertical de porcentajes
    for var_nombre, var_info in variables.items():
        col_original = var_info['columna']
        num_cats = len(var_info['categorias'])
        
        for i, cat in enumerate(var_info['categorias']):
            es_primera = (i == 0)
            es_ultima = (i == num_cats - 1)
            
            # Sumar los porcentajes verticalmente
            suma_porcentajes = 0
            for idx_opcion in range(len(opciones)):
                fila_anterior = fila_porcentajes - len(opciones) + idx_opcion
                valor_celda = ws.cell(row=fila_anterior, column=col_actual).value
                if valor_celda is not None:
                    if isinstance(valor_celda, (int, float)):
                        # Los valores están en formato decimal (0.13), multiplicar por 10000 para obtener centésimas
                        suma_porcentajes += int(valor_celda * 10000)
                    elif isinstance(valor_celda, str) and valor_celda != "---" and '%' in valor_celda:
                        suma_porcentajes += int(valor_celda.replace('%', '')) * 100
            
            if suma_porcentajes == 0:
                ws.cell(row=fila_porcentajes, column=col_actual, value="---")
            else:
                # Truncar a dos decimales sin redondear (almacenar como decimal con 2 decimales)
                suma_decimal = int(suma_porcentajes) / 10000
                ws.cell(row=fila_porcentajes, column=col_actual, value=suma_decimal)
                ws.cell(row=fila_porcentajes, column=col_actual).number_format = '0.00%'
            
            ws.cell(row=fila_porcentajes, column=col_actual).font = Font(bold=True)
            border = Border(
                left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),
                top=Side(style='thin', color='FFD0D0D0'),
                bottom=Side(style='medium')
            )
            ws.cell(row=fila_porcentajes, column=col_actual).border = border
            ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
            col_actual += 1
    
    # Asegurar bordes en última fila y última columna de porcentajes
    for col_idx in range(1, total_columnas + 1):
        cell = ws.cell(row=fila_porcentajes, column=col_idx)
        if cell.border:
            current_border = cell.border
            right_style = Side(style='medium') if col_idx == total_columnas else current_border.right
            new_border = Border(
                left=current_border.left,
                right=right_style,
                top=current_border.top,
                bottom=Side(style='medium')
            )
            cell.border = new_border
    
    # Asegurar que la última columna de todas las filas tenga right=medium
    for row in range(fila - 2, fila_porcentajes + 1):
        cell = ws.cell(row=row, column=total_columnas)
        if cell.border:
            current_border = cell.border
            new_border = Border(
                left=current_border.left,
                right=Side(style='medium'),
                top=current_border.top,
                bottom=current_border.bottom
            )
            cell.border = new_border
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 30
    for col_idx in range(2, total_columnas + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    print(f"  ✓ Hoja P{pregunta_num} generada exitosamente")

def generar_analisis_en_hoja_unica(ws, df, pregunta_num, pregunta_col, pregunta_nombre, tiene_combinaciones, fila_inicio):
    """
    Genera el análisis de una pregunta en una hoja existente, empezando desde fila_inicio.
    Retorna la siguiente fila disponible.
    """
    # Obtener opciones de la pregunta
    opciones = obtener_opciones_unicas(df, pregunta_col, tiene_combinaciones)
    
    if len(opciones) == 0:
        return fila_inicio
    
    # Preparar datos
    df_work = df.copy()
    
    # FILTRO ESPECIAL PARA P6: Solo incluir registros con "b. Contact Center" en P3
    if pregunta_col == 'P6 - Gestión Contact Center':
        p3_col = 'P3 - Medios SAT Utilizados'
        if p3_col in df_work.columns:
            # Filtrar solo registros que tienen "b. Contact Center" en P3
            mask_p3_contact = df_work[p3_col].astype(str).str.contains('b. Contact Center', na=False)
            df_work = df_work[mask_p3_contact].copy()
    
    # FILTRO ESPECIAL PARA P7: Solo incluir registros con "b. Contact Center" en P3
    if pregunta_col == 'P7 - Medio Contact Center':
        p3_col = 'P3 - Medios SAT Utilizados'
        if p3_col in df_work.columns:
            # Filtrar solo registros que tienen "b. Contact Center" en P3
            mask_p3_contact = df_work[p3_col].astype(str).str.contains('b. Contact Center', na=False)
            df_work = df_work[mask_p3_contact].copy()
    
    # FILTRO ESPECIAL PARA P8: Solo incluir registros con "a. Presencial" en P3
    if pregunta_col == 'P8 - Gestión Visita Presencial':
        p3_col = 'P3 - Medios SAT Utilizados'
        if p3_col in df_work.columns:
            # Filtrar solo registros que tienen "a. Presencial" en P3
            mask_p3_presencial = df_work[p3_col].astype(str).str.contains('a. Presencial', na=False)
            df_work = df_work[mask_p3_presencial].copy()
    
    df_work['Rango_Edad'] = df_work['P36 - Edad'].apply(crear_rango_edad)
    df_work['Region_Oficina'] = df_work['P44 - Oficina/Agencia/Delegación'].apply(obtener_region_oficina)
    df_work['Region_Aduana'] = df_work['P44 - Aduana'].apply(obtener_region_aduana)
    
    if tiene_combinaciones:
        df_work[f'{pregunta_col}_norm'] = df_work[pregunta_col].apply(
            lambda x: normalizar_combinaciones(x, opciones)
        )
    
    # Definir estilos
    thin_side = Side(style='thin', color='FFD0D0D0')
    medium_side = Side(style='medium')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    fill_fila1 = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    fill_header = PatternFill(start_color='FFE7E6E6', end_color='FFE7E6E6', fill_type='solid')
    
    # Definir variables de cruce
    col = 1
    variables = {
        'P37 Género': {
            'columna': 'P37 - Género',
            'categorias': ['H', 'M', 'No deseo responder'],
            'col_inicio': col + 2
        },
        'P3 Medios SAT utilizados': {
            'columna': 'P3 - Medios SAT Utilizados',
            'categorias': ['a. Presencial', 'b. Contact Center', 'c. Servicios Electrónicos'],
            'col_inicio': col + 5,
            'usa_contains': True  # Marcar que usa str.contains() para contar
        },
        'Rango de edad': {
            'columna': 'Rango_Edad',
            'categorias': ['18 - 25', '26 - 35', '36 - 45', '46 - 60', 'Más de 61'],
            'col_inicio': col + 8
        },
        'P40 Nivel académico': {
            'columna': 'P40 - Nivel Académico',
            'categorias': [
                'a. Ninguno', 'b. Primaria incompleta', 'c. Primaria completa',
                'd. Secundaria incompleta (1ro a 3ro básico)', 'e. Secundaria Completa (1ro a 3ro básico)',
                'f. Diversificado incompleto', 'g. Diversificado completo', 'h. Técnico',
                'i. Universidad incompleta', 'j. Universidad Completa', 'k. Maestría / Posgrado'
            ],
            'col_inicio': col + 13
        },
        'P39 Idiomas': {
            'columna': 'P39 - Idiomas',
            'categorias': [
                'a. Achi', 'b. Qánjob\'al', 'c. Q\'eqchi', 'd. Akateco', 'e. Kaqchikel',
                'f. Sakapulteko', 'h. Kiché', 'i. Sipakapense', 'k. Mam', 'n. Mopan',
                'p. Ixil', 'q. Poqomam', 's. Jakalteco', 't. Poqomchi', 'u. Ninguno',
                'v. Inglés', 'w. Otro'
            ],
            'col_inicio': col + 24
        },
        'P44 Oficina/Agencia/Delegación': {
            'columna': 'P44 - Oficina/Agencia/Delegación',
            'categorias': [
                'Alta Verapaz', 'Baja Verapaz', 'Chimaltenango', 'Chiquimula', 'El Progreso',
                'Escuintla', 'Guatemala', 'Huehuetenango', 'Izabal', 'Jalapa', 'Jutiapa',
                'Petén', 'Quetzaltenango', 'Quiché', 'Retalhuleu', 'Sacatepéquez', 'San Marcos',
                'Santa Rosa', 'Sololá', 'Suchitepéquez', 'Totonicapán', 'Zacapa'
            ],
            'col_inicio': col + 41
        },
        'P44.1 Aduana': {
            'columna': 'P44 - Aduana',
            'categorias': [
                'Central Guatemala', 'El Carmen', 'Integrada Corinto', 'Integrada El Florido',
                'La Mesilla', 'Puerto Barrios Almacenadora Pelícano, S.A -ALPELSA', 'Puerto Quetzal',
                'San Cristóbal', 'Santo Tomás de Castilla Zona Libre de Industria y Comercio -ZOLIC-',
                'Tikal', 'Valle Nuevo'
            ],
            'col_inicio': col + 63
        },
        'P9 Personería': {
            'columna': 'P9 - Personería',
            'categorias': [
                'a. Contribuyente/Propietario.', 'b. Representante Legal', 'c. Abogado y Notario',
                'd. Mandatario', 'e. Contador/auxiliar', 'f. Contador Público y Auditor',
                'g. Gestor Tributario', 'h. Importador', 'i. Exportador', 'j. Asistente de Agente',
                'k. Auxiliar Gestor Tributario', 'm. Consolidador/Descons.', 'n. Transportista Ad',
                'p. Mensajero', 'r. Otro'
            ],
            'col_inicio': col + 74
        },
        'P38 Etnia': {
            'columna': 'P38 - Etnia',
            'categorias': ['Garifuna', 'Ladino', 'Maya', 'Otro', 'Xinca'],
            'col_inicio': col + 89
        },
        'Oficina/Agencia/Delegación': {
            'columna': 'Region_Oficina',
            'categorias': ['Central', 'Occidente', 'Sur', 'Nororiente'],
            'col_inicio': col + 94
        },
        'Aduana': {
            'columna': 'Region_Aduana',
            'categorias': ['Central', 'Occidente', 'Sur', 'Nororiente'],
            'col_inicio': col + 98
        }
    }
    
    # Calcular total de columnas
    total_columnas = 2
    for var_nombre, var_info in variables.items():
        total_columnas += len(var_info['categorias'])
    
    fila = fila_inicio
    
    # Fila: Título de la pregunta
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=total_columnas)
    ws.cell(row=fila, column=1, value=pregunta_nombre)
    ws.cell(row=fila, column=1).font = Font(bold=True, size=14)
    ws.cell(row=fila, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=fila, column=1).fill = fill_fila1
    ws.cell(row=fila, column=1).border = Border(
        left=medium_side,
        right=medium_side,
        top=medium_side,
        bottom=thin_side
    )
    fila += 1
    
    # Fila: Vacía con fondo gris
    for col_idx in range(1, total_columnas + 1):
        cell = ws.cell(row=fila, column=col_idx)
        cell.fill = fill_fila1
        cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    fila += 1
    
    # Fila: Encabezados principales
    col_actual = 1
    ws.cell(row=fila, column=col_actual, value='')
    ws.cell(row=fila, column=col_actual).fill = fill_header
    ws.cell(row=fila, column=col_actual).border = Border(left=medium_side, right=medium_side, top=thin_side, bottom=thin_side)
    col_actual += 1
    
    ws.cell(row=fila, column=col_actual, value='TOTAL')
    ws.cell(row=fila, column=col_actual).fill = fill_header
    ws.cell(row=fila, column=col_actual).border = Border(left=medium_side, right=medium_side, top=thin_side, bottom=thin_side)
    ws.cell(row=fila, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    col_actual += 1
    
    for var_nombre, var_info in variables.items():
        num_cols = len(var_info['categorias'])
        inicio = col_actual
        fin = col_actual + num_cols - 1
        
        ws.merge_cells(start_row=fila, start_column=inicio, end_row=fila, end_column=fin)
        ws.cell(row=fila, column=inicio, value=var_nombre)
        ws.cell(row=fila, column=inicio).fill = fill_header
        ws.cell(row=fila, column=inicio).border = Border(left=medium_side, right=medium_side, top=thin_side, bottom=thin_side)
        ws.cell(row=fila, column=inicio).font = Font(bold=True)
        ws.cell(row=fila, column=inicio).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        col_actual = fin + 1
    fila += 1
    
    # Fila: Sub-encabezados
    col_actual = 1
    ws.cell(row=fila, column=col_actual, value='')
    ws.cell(row=fila, column=col_actual).border = thin_border
    col_actual += 1
    
    ws.cell(row=fila, column=col_actual, value='')
    ws.cell(row=fila, column=col_actual).border = Border(
        left=medium_side, 
        right=medium_side, 
        top=Side(style='medium'), 
        bottom=Side(style='thin', color='FFD0D0D0')
    )
    col_actual += 1
    
    for var_nombre, var_info in variables.items():
        num_cats = len(var_info['categorias'])
        for i, cat in enumerate(var_info['categorias']):
            es_primera = (i == 0)
            es_ultima = (i == num_cats - 1)
            ws.cell(row=fila, column=col_actual, value=cat)
            ws.cell(row=fila, column=col_actual).font = Font(bold=True)
            border = Border(
                left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),
                top=Side(style='medium'),
                bottom=Side(style='thin', color='FFD0D0D0')
            )
            ws.cell(row=fila, column=col_actual).border = border
            ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            col_actual += 1
    fila += 1
    
    # Filas de datos
    if tiene_combinaciones:
        df_pregunta = df_work[df_work[pregunta_col].notna()].copy()
    else:
        df_pregunta = df_work[df_work[pregunta_col].notna()].copy()
    
    for idx_opcion, opcion in enumerate(opciones):
        col_actual = 1
        
        # Nombre de la fila
        ws.cell(row=fila, column=col_actual, value=opcion)
        border = Border(
            left=Side(style='thin', color='FFD0D0D0'),
            right=Side(style='thin', color='FFD0D0D0'),
            top=Side(style='medium' if idx_opcion == 0 else 'thin', color='FFD0D0D0'),
            bottom=Side(style='thin', color='FFD0D0D0')
        )
        ws.cell(row=fila, column=col_actual).border = border
        ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='left', vertical='center')
        col_actual += 1
        
        # TOTAL
        if tiene_combinaciones:
            total = len(df_work[df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)])
        else:
            total = len(df_work[df_work[pregunta_col] == opcion])
        
        ws.cell(row=fila, column=col_actual, value=total)
        border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium' if idx_opcion == 0 else 'thin', color='FFD0D0D0'),
            bottom=Side(style='thin', color='FFD0D0D0')
        )
        ws.cell(row=fila, column=col_actual).border = border
        ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
        col_actual += 1
        
        # Datos por variable
        for var_nombre, var_info in variables.items():
            col_original = var_info['columna']
            num_cats = len(var_info['categorias'])
            
            for i, cat in enumerate(var_info['categorias']):
                es_primera = (i == 0)
                es_ultima = (i == num_cats - 1)
                
                # Contar intersección
                if tiene_combinaciones:
                    if col_original == 'Rango_Edad':
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original] == cat)
                        ])
                    elif col_original == 'P39 - Idiomas':
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original].astype(str).str.contains(cat, na=False))
                        ])
                    elif col_original == 'P3 - Medios SAT Utilizados':
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original].astype(str).str.contains(cat, na=False))
                        ])
                    elif col_original == 'Region_Oficina' or col_original == 'Region_Aduana':
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original] == cat)
                        ])
                    else:
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original] == cat)
                        ])
                else:
                    if col_original == 'Rango_Edad':
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original] == cat)
                        ])
                    elif col_original == 'P39 - Idiomas':
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original].astype(str).str.contains(cat, na=False))
                        ])
                    elif col_original == 'P3 - Medios SAT Utilizados':
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original].astype(str).str.contains(cat, na=False))
                        ])
                    elif col_original == 'Region_Oficina' or col_original == 'Region_Aduana':
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original] == cat)
                        ])
                    else:
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original] == cat)
                        ])
                
                ws.cell(row=fila, column=col_actual, value=count)
                border = Border(
                    left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                    right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),
                    top=Side(style='medium' if idx_opcion == 0 else 'thin', color='FFD0D0D0'),
                    bottom=Side(style='thin', color='FFD0D0D0')
                )
                ws.cell(row=fila, column=col_actual).border = border
                ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
                col_actual += 1
        
        fila += 1
    
    # Asegurar bordes right=medium en última columna
    for row in range(fila_inicio, fila + 1):
        cell = ws.cell(row=row, column=total_columnas)
        if cell.border:
            current_border = cell.border
            new_border = Border(
                left=current_border.left,
                right=Side(style='medium'),
                top=current_border.top,
                bottom=current_border.bottom
            )
            cell.border = new_border
    
    # Fila TOTAL
    col_actual = 1
    ws.cell(row=fila, column=col_actual, value='TOTAL')
    ws.cell(row=fila, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila, column=col_actual).border = Border(
        left=Side(style='thin', color='FFD0D0D0'),
        right=Side(style='thin', color='FFD0D0D0'),
        top=Side(style='thin', color='FFD0D0D0'),
        bottom=Side(style='medium')
    )
    ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
    col_actual += 1
    
    # TOTAL general
    total_general = len(df_pregunta)
    ws.cell(row=fila, column=col_actual, value=total_general)
    ws.cell(row=fila, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila, column=col_actual).border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='thin', color='FFD0D0D0'),
        bottom=Side(style='medium')
    )
    ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
    col_actual += 1
    
    # Totales por categoría
    for var_nombre, var_info in variables.items():
        col_original = var_info['columna']
        num_cats = len(var_info['categorias'])
        
        for i, cat in enumerate(var_info['categorias']):
            es_primera = (i == 0)
            es_ultima = (i == num_cats - 1)
            
            if col_original == 'Rango_Edad':
                total_cat = len(df_pregunta[df_pregunta[col_original] == cat])
            elif col_original == 'P39 - Idiomas':
                total_cat = len(df_pregunta[df_pregunta[col_original].astype(str).str.contains(cat, na=False)])
            elif col_original == 'P3 - Medios SAT Utilizados':
                total_cat = len(df_pregunta[df_pregunta[col_original].astype(str).str.contains(cat, na=False)])
            elif col_original == 'Region_Oficina' or col_original == 'Region_Aduana':
                total_cat = len(df_pregunta[df_pregunta[col_original] == cat])
            else:
                total_cat = len(df_pregunta[df_pregunta[col_original] == cat])
            
            ws.cell(row=fila, column=col_actual, value=total_cat)
            ws.cell(row=fila, column=col_actual).font = Font(bold=True)
            border = Border(
                left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),
                top=Side(style='thin', color='FFD0D0D0'),
                bottom=Side(style='medium')
            )
            ws.cell(row=fila, column=col_actual).border = border
            ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
            col_actual += 1
    
    # Asegurar bordes en última fila
    for col_idx in range(1, total_columnas + 1):
        cell = ws.cell(row=fila, column=col_idx)
        if cell.border:
            current_border = cell.border
            right_style = Side(style='medium') if col_idx == total_columnas else current_border.right
            new_border = Border(
                left=current_border.left,
                right=right_style,
                top=current_border.top,
                bottom=Side(style='medium')
            )
            cell.border = new_border
    
    # Agregar dos filas vacías
    fila += 2
    
    # TABLA DE PORCENTAJES
    fila_porcentajes = fila
    
    # Fila de encabezados principales
    col_actual = 1
    ws.cell(row=fila_porcentajes, column=col_actual, value='')
    ws.cell(row=fila_porcentajes, column=col_actual).fill = fill_header
    ws.cell(row=fila_porcentajes, column=col_actual).border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=thin_side)
    col_actual += 1
    
    ws.cell(row=fila_porcentajes, column=col_actual, value='TOTAL')
    ws.cell(row=fila_porcentajes, column=col_actual).fill = fill_header
    ws.cell(row=fila_porcentajes, column=col_actual).border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=thin_side)
    ws.cell(row=fila_porcentajes, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    col_actual += 1
    
    for var_nombre, var_info in variables.items():
        num_cols = len(var_info['categorias'])
        inicio = col_actual
        fin = col_actual + num_cols - 1
        
        ws.merge_cells(start_row=fila_porcentajes, start_column=inicio, end_row=fila_porcentajes, end_column=fin)
        ws.cell(row=fila_porcentajes, column=inicio, value=var_nombre)
        ws.cell(row=fila_porcentajes, column=inicio).fill = fill_header
        ws.cell(row=fila_porcentajes, column=inicio).border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=thin_side)
        ws.cell(row=fila_porcentajes, column=inicio).font = Font(bold=True)
        ws.cell(row=fila_porcentajes, column=inicio).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        col_actual = fin + 1
    
    # Fila de sub-encabezados
    fila_porcentajes += 1
    col_actual = 1
    ws.cell(row=fila_porcentajes, column=col_actual, value='')
    ws.cell(row=fila_porcentajes, column=col_actual).border = thin_border
    col_actual += 1
    
    ws.cell(row=fila_porcentajes, column=col_actual, value='')
    ws.cell(row=fila_porcentajes, column=col_actual).border = Border(
        left=medium_side, 
        right=medium_side, 
        top=Side(style='medium'), 
        bottom=Side(style='thin', color='FFD0D0D0')
    )
    col_actual += 1
    
    for var_nombre, var_info in variables.items():
        num_cats = len(var_info['categorias'])
        for i, cat in enumerate(var_info['categorias']):
            es_primera = (i == 0)
            es_ultima = (i == num_cats - 1)
            ws.cell(row=fila_porcentajes, column=col_actual, value=cat)
            ws.cell(row=fila_porcentajes, column=col_actual).font = Font(bold=True)
            border = Border(
                left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),
                top=Side(style='medium'),
                bottom=Side(style='thin', color='FFD0D0D0')
            )
            ws.cell(row=fila_porcentajes, column=col_actual).border = border
            ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            col_actual += 1
    
    # Filas de datos con porcentajes
    fila_porcentajes += 1
    for idx_opcion, opcion in enumerate(opciones):
        col_actual = 1
        
        # Nombre de la fila
        ws.cell(row=fila_porcentajes, column=col_actual, value=opcion)
        border = Border(
            left=Side(style='thin', color='FFD0D0D0'),
            right=Side(style='thin', color='FFD0D0D0'),
            top=Side(style='medium' if idx_opcion == 0 else 'thin', color='FFD0D0D0'),
            bottom=Side(style='thin', color='FFD0D0D0')
        )
        ws.cell(row=fila_porcentajes, column=col_actual).border = border
        ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='left', vertical='center')
        col_actual += 1
        
        # TOTAL - calcular porcentaje
        if tiene_combinaciones:
            total_absoluto = len(df_work[df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)])
        else:
            total_absoluto = len(df_work[df_work[pregunta_col] == opcion])
        
        porcentaje_total = (total_absoluto / total_general * 100) if total_general > 0 else 0
        # Truncar a dos decimales sin redondear (almacenar como decimal)
        # Ejemplo: 13.456% -> int(13.456 * 100) = 1345 -> 1345/10000 = 0.1345 (representa 13.45%)
        porcentaje_decimal = int(porcentaje_total * 100) / 10000
        
        if porcentaje_decimal == 0:
            ws.cell(row=fila_porcentajes, column=col_actual, value="---")
        else:
            ws.cell(row=fila_porcentajes, column=col_actual, value=porcentaje_decimal)
            ws.cell(row=fila_porcentajes, column=col_actual).number_format = '0.00%'
        
        border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium' if idx_opcion == 0 else 'thin', color='FFD0D0D0'),
            bottom=Side(style='thin', color='FFD0D0D0')
        )
        ws.cell(row=fila_porcentajes, column=col_actual).border = border
        ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
        col_actual += 1
        
        # Datos por variable (porcentajes)
        for var_nombre, var_info in variables.items():
            col_original = var_info['columna']
            num_cats = len(var_info['categorias'])
            
            for i, cat in enumerate(var_info['categorias']):
                es_primera = (i == 0)
                es_ultima = (i == num_cats - 1)
                
                # Contar intersección
                if tiene_combinaciones:
                    if col_original == 'Rango_Edad':
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original] == cat)
                        ])
                    elif col_original == 'P39 - Idiomas':
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original].astype(str).str.contains(cat, na=False))
                        ])
                    elif col_original == 'P3 - Medios SAT Utilizados':
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original].astype(str).str.contains(cat, na=False))
                        ])
                    elif col_original == 'Region_Oficina' or col_original == 'Region_Aduana':
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original] == cat)
                        ])
                    else:
                        count = len(df_work[
                            (df_work[pregunta_col].astype(str).str.contains(opcion, na=False, regex=False)) & 
                            (df_work[col_original] == cat)
                        ])
                else:
                    if col_original == 'Rango_Edad':
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original] == cat)
                        ])
                    elif col_original == 'P39 - Idiomas':
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original].astype(str).str.contains(cat, na=False))
                        ])
                    elif col_original == 'P3 - Medios SAT Utilizados':
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original].astype(str).str.contains(cat, na=False))
                        ])
                    elif col_original == 'Region_Oficina' or col_original == 'Region_Aduana':
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original] == cat)
                        ])
                    else:
                        count = len(df_work[
                            (df_work[pregunta_col] == opcion) & 
                            (df_work[col_original] == cat)
                        ])
                
                # Calcular porcentaje VERTICAL (sobre el total de esa categoría)
                if col_original == 'Rango_Edad':
                    total_categoria = len(df_pregunta[df_pregunta[col_original] == cat])
                elif col_original == 'P39 - Idiomas':
                    total_categoria = len(df_pregunta[df_pregunta[col_original].astype(str).str.contains(cat, na=False)])
                elif col_original == 'P3 - Medios SAT Utilizados':
                    total_categoria = len(df_pregunta[df_pregunta[col_original].astype(str).str.contains(cat, na=False)])
                elif col_original == 'Region_Oficina' or col_original == 'Region_Aduana':
                    total_categoria = len(df_pregunta[df_pregunta[col_original] == cat])
                else:
                    total_categoria = len(df_pregunta[df_pregunta[col_original] == cat])
                
                porcentaje = (count / total_categoria * 100) if total_categoria > 0 else 0
                # Truncar a dos decimales sin redondear (almacenar como decimal con 2 decimales)
                porcentaje_decimal = int(porcentaje * 100) / 10000
                
                if porcentaje_decimal == 0:
                    ws.cell(row=fila_porcentajes, column=col_actual, value="---")
                else:
                    ws.cell(row=fila_porcentajes, column=col_actual, value=porcentaje_decimal)
                    ws.cell(row=fila_porcentajes, column=col_actual).number_format = '0.00%'
                
                border = Border(
                    left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                    right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),
                    top=Side(style='medium' if idx_opcion == 0 else 'thin', color='FFD0D0D0'),
                    bottom=Side(style='thin', color='FFD0D0D0')
                )
                ws.cell(row=fila_porcentajes, column=col_actual).border = border
                ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
                col_actual += 1
        
        fila_porcentajes += 1
    
    # Fila TOTAL de porcentajes
    col_actual = 1
    ws.cell(row=fila_porcentajes, column=col_actual, value='TOTAL')
    ws.cell(row=fila_porcentajes, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila_porcentajes, column=col_actual).border = Border(
        left=Side(style='thin', color='FFD0D0D0'),
        right=Side(style='thin', color='FFD0D0D0'),
        top=Side(style='thin', color='FFD0D0D0'),
        bottom=Side(style='medium')
    )
    ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
    col_actual += 1
    
    # TOTAL general (100%)
    ws.cell(row=fila_porcentajes, column=col_actual, value=1.0)
    ws.cell(row=fila_porcentajes, column=col_actual).number_format = '0%'
    ws.cell(row=fila_porcentajes, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila_porcentajes, column=col_actual).border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='thin', color='FFD0D0D0'),
        bottom=Side(style='medium')
    )
    ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
    col_actual += 1
    
    # Totales por categoría (suma vertical de porcentajes)
    for var_nombre, var_info in variables.items():
        col_original = var_info['columna']
        num_cats = len(var_info['categorias'])
        
        for i, cat in enumerate(var_info['categorias']):
            es_primera = (i == 0)
            es_ultima = (i == num_cats - 1)
            
            # Sumar porcentajes verticalmente
            suma_porcentajes = 0
            for idx_opcion, opcion in enumerate(opciones):
                fila_opcion = fila_porcentajes - len(opciones) + idx_opcion
                cell = ws.cell(row=fila_opcion, column=col_actual)
                if cell.value != "---" and cell.value is not None:
                    suma_porcentajes += cell.value if isinstance(cell.value, (int, float)) else 0
            
            if suma_porcentajes == 0:
                ws.cell(row=fila_porcentajes, column=col_actual, value="---")
            else:
                ws.cell(row=fila_porcentajes, column=col_actual, value=suma_porcentajes)
                ws.cell(row=fila_porcentajes, column=col_actual).number_format = '0%'
            
            ws.cell(row=fila_porcentajes, column=col_actual).font = Font(bold=True)
            border = Border(
                left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),
                top=Side(style='thin', color='FFD0D0D0'),
                bottom=Side(style='medium')
            )
            ws.cell(row=fila_porcentajes, column=col_actual).border = border
            ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
            col_actual += 1
    
    # Asegurar bordes en última fila de porcentajes
    for col_idx in range(1, total_columnas + 1):
        cell = ws.cell(row=fila_porcentajes, column=col_idx)
        if cell.border:
            current_border = cell.border
            right_style = Side(style='medium') if col_idx == total_columnas else current_border.right
            new_border = Border(
                left=current_border.left,
                right=right_style,
                top=current_border.top,
                bottom=Side(style='medium')
            )
            cell.border = new_border
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 30
    for col_idx in range(2, total_columnas + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    return fila_porcentajes + 1

def generar_todos_analisis(archivo_entrada='V3.xlsx', archivo_salida='Todos-Cruzado.xlsx'):
    """
    Función principal que genera análisis cruzado de todas las preguntas desde P3.
    """
    print(f"Leyendo archivo: {archivo_entrada}")
    
    if not os.path.exists(archivo_entrada):
        print(f"ERROR: No se encontró el archivo {archivo_entrada}")
        sys.exit(1)
    
    try:
        df = pd.read_excel(archivo_entrada)
        print(f"Archivo leído exitosamente. Total de registros: {len(df)}")
    except Exception as e:
        print(f"ERROR al leer el archivo: {e}")
        sys.exit(1)
    
    # Crear workbook
    wb = Workbook()
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # Obtener TODAS las preguntas desde P3 (incluyendo todas las variantes)
    columnas = df.columns.tolist()
    preguntas = []
    
    for col in columnas:
        if col.startswith('P') and ' - ' in col:
            # Extraer número de pregunta (puede ser P3, P11.1, P12.1, etc.)
            num_pregunta_str = col.split(' - ')[0].replace('P', '')
            try:
                # Intentar convertir a float para manejar P11.1, P12.1, etc.
                num = float(num_pregunta_str)
                if num >= 3:
                    # Incluir TODAS las variantes (no solo la primera)
                    preguntas.append((num, col, num_pregunta_str))
            except:
                pass
    
    # Ordenar por número
    preguntas.sort(key=lambda x: x[0])
    
    print(f"\n{'='*80}")
    print(f"PREGUNTAS ENCONTRADAS: {len(preguntas)}")
    print(f"{'='*80}")
    
    # ============================================================================
    # VERSIÓN 1: CON PESTAÑAS (cada pregunta en su propia hoja)
    # ============================================================================
    print(f"\n{'='*80}")
    print("GENERANDO VERSIÓN CON PESTAÑAS")
    print(f"{'='*80}")
    
    wb_pestanas = Workbook()
    wb_pestanas.remove(wb_pestanas.active)
    
    for pregunta_num, pregunta_col, num_str in preguntas:
        pregunta_nombre = pregunta_col
        
        # Detectar si tiene combinaciones múltiples
        tiene_combinaciones = detectar_combinaciones_multiples(df, pregunta_col)
        
        # Generar hoja
        try:
            generar_hoja_pregunta(wb_pestanas, df, num_str, pregunta_col, pregunta_nombre, tiene_combinaciones)
        except Exception as e:
            print(f"  ✗ Error al procesar {pregunta_nombre}: {e}")
            import traceback
            traceback.print_exc()
    
    # Guardar archivo con pestañas
    archivo_pestanas = archivo_salida.replace('.xlsx', '-Pestanas.xlsx')
    print(f"\n{'='*80}")
    print(f"Guardando archivo con pestañas: {archivo_pestanas}")
    try:
        wb_pestanas.save(archivo_pestanas)
        print(f"✓ Archivo generado exitosamente: {archivo_pestanas}")
        print(f"  Total de hojas generadas: {len(wb_pestanas.worksheets)}")
    except Exception as e:
        print(f"ERROR al guardar el archivo: {e}")
        sys.exit(1)
    
    # ============================================================================
    # VERSIÓN 2: UNA SOLA HOJA (todas las preguntas en la misma hoja)
    # ============================================================================
    print(f"\n{'='*80}")
    print("GENERANDO VERSIÓN EN UNA SOLA HOJA")
    print(f"{'='*80}")
    
    wb_una_hoja = Workbook()
    ws_unica = wb_una_hoja.active
    ws_unica.title = "Todos los Análisis"
    
    fila_actual = 1
    
    for pregunta_num, pregunta_col, num_str in preguntas:
        pregunta_nombre = pregunta_col
        
        # Detectar si tiene combinaciones múltiples
        tiene_combinaciones = detectar_combinaciones_multiples(df, pregunta_col)
        
        # Generar análisis en la misma hoja
        try:
            fila_actual = generar_analisis_en_hoja_unica(
                ws_unica, df, num_str, pregunta_col, pregunta_nombre, 
                tiene_combinaciones, fila_actual
            )
            # Agregar 3 filas vacías entre preguntas
            fila_actual += 3
        except Exception as e:
            print(f"  ✗ Error al procesar {pregunta_nombre}: {e}")
            import traceback
            traceback.print_exc()
    
    # Guardar archivo en una sola hoja
    archivo_una_hoja = archivo_salida.replace('.xlsx', '-UnaHoja.xlsx')
    print(f"\n{'='*80}")
    print(f"Guardando archivo en una sola hoja: {archivo_una_hoja}")
    try:
        wb_una_hoja.save(archivo_una_hoja)
        print(f"✓ Archivo generado exitosamente: {archivo_una_hoja}")
        print(f"  Total de filas generadas: {fila_actual}")
    except Exception as e:
        print(f"ERROR al guardar el archivo: {e}")
        sys.exit(1)

if __name__ == "__main__":
    archivo_entrada = sys.argv[1] if len(sys.argv) > 1 else 'V3.xlsx'
    archivo_salida = sys.argv[2] if len(sys.argv) > 2 else 'Todos-Cruzado.xlsx'
    
    print("=" * 80)
    print("GENERADOR DE ANÁLISIS CRUZADO - TODAS LAS PREGUNTAS")
    print("=" * 80)
    print()
    
    generar_todos_analisis(archivo_entrada, archivo_salida)
    
    print()
    print("=" * 80)
    print("Proceso completado exitosamente")
    print("=" * 80)

