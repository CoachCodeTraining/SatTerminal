#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para generar análisis cruzado de la pregunta P4 (Servicio Electrónico)
con todas las variables demográficas y de clasificación.

Autor: Generado automáticamente
Fecha: 2025
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys

def normalizar_p4(valor):
    """
    Normaliza los valores de P4. P4 no tiene combinaciones múltiples,
    solo valores individuales.
    """
    if pd.isna(valor):
        return None
    valor_str = str(valor).strip()
    return valor_str if valor_str else None

def crear_rango_edad(edad):
    """
    Crea rangos de edad a partir de la edad numérica.
    """
    if pd.isna(edad):
        return None
    try:
        edad_num = int(float(edad))
        if edad_num <= 25:
            return '18 - 25'
        elif edad_num <= 35:
            return '26 - 35'
        elif edad_num <= 45:
            return '36 - 45'
        elif edad_num <= 60:
            return '46 - 60'
        else:
            return 'Más de 61'
    except:
        return None

def obtener_region_oficina(oficina):
    """
    Agrupa las oficinas/agencias/delegaciones por región.
    """
    if pd.isna(oficina):
        return None
    
    oficina_str = str(oficina).strip()
    
    # Central
    central = ['Chimaltenango', 'El Progreso', 'Guatemala', 'Sacatepéquez']
    if oficina_str in central:
        return 'Central'
    
    # Occidente
    occidente = ['Huehuetenango', 'Quetzaltenango', 'Quiché', 'San Marcos', 'Sololá', 'Totonicapán']
    if oficina_str in occidente:
        return 'Occidente'
    
    # Sur
    sur = ['Escuintla', 'Jutiapa', 'Retalhuleu', 'Santa Rosa', 'Suchitepéquez']
    if oficina_str in sur:
        return 'Sur'
    
    # Nororiente
    nororiente = ['Alta Verapaz', 'Baja Verapaz', 'Chiquimula', 'Izabal', 'Jalapa', 'Petén', 'Zacapa']
    if oficina_str in nororiente:
        return 'Nororiente'
    
    return None

def obtener_region_aduana(aduana):
    """
    Agrupa las aduanas por región.
    """
    if pd.isna(aduana):
        return None
    
    aduana_str = str(aduana).strip()
    
    # Central
    if 'Central Guatemala' in aduana_str:
        return 'Central'
    
    # Occidente
    if 'El Carmen' in aduana_str or 'La Mesilla' in aduana_str:
        return 'Occidente'
    
    # Sur
    if 'San Cristóbal' in aduana_str or 'Valle Nuevo' in aduana_str or 'Puerto Quetzal' in aduana_str:
        return 'Sur'
    
    # Nororiente
    if 'Integrada Corinto' in aduana_str or 'Integrada El Florido' in aduana_str or 'Puerto Barrios' in aduana_str or 'Santo Tomás' in aduana_str or 'Tikal' in aduana_str:
        return 'Nororiente'
    
    return None

def aplicar_estilos_bordes(ws, fila, col, es_primera_fila=False, es_ultima_fila=False, 
                           es_primera_col=False, es_ultima_col=False, es_encabezado=False):
    """
    Aplica los estilos de bordes según la posición de la celda.
    """
    thin_side = Side(style='thin', color='FFD0D0D0')
    medium_side = Side(style='medium')
    
    left_style = medium_side if es_primera_col else thin_side
    right_style = medium_side if es_ultima_col else thin_side
    top_style = medium_side if es_primera_fila else thin_side
    bottom_style = medium_side if es_ultima_fila else thin_side
    
    border = Border(
        left=left_style,
        right=right_style,
        top=top_style,
        bottom=bottom_style
    )
    
    ws.cell(row=fila, column=col).border = border
    return border

def generar_analisis_cruzado(archivo_entrada='V3.xlsx', archivo_salida='Analisis_Cruzado_P4.xlsx'):
    """
    Función principal que genera el análisis cruzado de P4.
    
    Args:
        archivo_entrada: Nombre del archivo Excel de entrada (default: V3.xlsx)
        archivo_salida: Nombre del archivo Excel de salida (default: Analisis_Cruzado_P4.xlsx)
    """
    print(f"Leyendo archivo: {archivo_entrada}")
    
    # Verificar que el archivo existe
    if not os.path.exists(archivo_entrada):
        print(f"ERROR: No se encontró el archivo {archivo_entrada}")
        sys.exit(1)
    
    # Leer el archivo
    try:
        df = pd.read_excel(archivo_entrada)
        print(f"Archivo leído exitosamente. Total de registros: {len(df)}")
    except Exception as e:
        print(f"ERROR al leer el archivo: {e}")
        sys.exit(1)
    
    # P4 no requiere normalización (no tiene combinaciones múltiples)
    print("Procesando valores de P4...")
    
    # Crear rangos de edad
    print("Creando rangos de edad...")
    df['Rango_Edad'] = df['P36 - Edad'].apply(crear_rango_edad)
    
    # Crear regiones para Oficina/Agencia/Delegación
    print("Creando regiones de Oficina/Agencia/Delegación...")
    df['Region_Oficina'] = df['P44 - Oficina/Agencia/Delegación'].apply(obtener_region_oficina)
    
    # Crear regiones para Aduana
    print("Creando regiones de Aduana...")
    df['Region_Aduana'] = df['P44 - Aduana'].apply(obtener_region_aduana)
    
    # Crear nuevo workbook
    print("Creando estructura del archivo Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "P4"
    
    # Definir estilos
    thin_side = Side(style='thin', color='FFD0D0D0')
    medium_side = Side(style='medium')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    fill_fila1 = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    fill_header = PatternFill(start_color='FFE7E6E6', end_color='FFE7E6E6', fill_type='solid')
    
    # Fila 0: Título de la pregunta (se ajustará después de calcular columnas)
    print("Agregando título de la pregunta...")
    ws.cell(row=1, column=1, value='P4 - Servicio Electrónico')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')  # Alineado a la izquierda
    ws.cell(row=1, column=1).fill = fill_fila1
    
    # Fila 1: Vacía con fondo gris (se ajustará después de calcular total_columnas)
    print("Configurando formato de la primera fila...")
    
    # Fila 3: Encabezados principales
    print("Creando encabezados principales...")
    col = 1
    ws.cell(row=3, column=col, value='')
    ws.cell(row=3, column=col).fill = fill_header
    ws.cell(row=3, column=col).border = Border(left=medium_side, right=medium_side, top=thin_side, bottom=thin_side)
    col += 1
    
    # TOTAL
    ws.cell(row=3, column=col, value='TOTAL')
    ws.cell(row=3, column=col).fill = fill_header
    ws.cell(row=3, column=col).border = Border(left=medium_side, right=medium_side, top=thin_side, bottom=thin_side)
    ws.cell(row=3, column=col).font = Font(bold=True)
    ws.cell(row=3, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    col += 1
    
    # Definir todas las variables y sus categorías
    variables = {
        'P37 Género': {
            'columna': 'P37 - Género',
            'categorias': ['H', 'M', 'No deseo responder'],
            'col_inicio': col
        },
        'Rango de edad': {
            'columna': 'Rango_Edad',
            'categorias': ['18 - 25', '26 - 35', '36 - 45', '46 - 60', 'Más de 61'],
            'col_inicio': col + 3
        },
        'P40 Nivel académico': {
            'columna': 'P40 - Nivel Académico',
            'categorias': [
                'a. Ninguno', 'b. Primaria incompleta', 'c. Primaria completa',
                'd. Secundaria incompleta (1ro a 3ro básico)', 'e. Secundaria Completa (1ro a 3ro básico)',
                'f. Diversificado incompleto', 'g. Diversificado completo', 'h. Técnico',
                'i. Universidad incompleta', 'j. Universidad Completa', 'k. Maestría / Posgrado'
            ],
            'col_inicio': col + 8
        },
        'P39 Idiomas': {
            'columna': 'P39 - Idiomas',
            'categorias': [
                'a. Achi', 'b. Qánjob\'al', 'c. Q\'eqchi', 'd. Akateco', 'e. Kaqchikel',
                'f. Sakapulteko', 'h. Kiché', 'i. Sipakapense', 'k. Mam', 'n. Mopan',
                'p. Ixil', 'q. Poqomam', 's. Jakalteco', 't. Poqomchi', 'u. Ninguno',
                'v. Inglés', 'w. Otro'
            ],
            'col_inicio': col + 19
        },
        'P44 Oficina/Agencia/Delegación': {
            'columna': 'P44 - Oficina/Agencia/Delegación',
            'categorias': [
                'Alta Verapaz', 'Baja Verapaz', 'Chimaltenango', 'Chiquimula', 'El Progreso',
                'Escuintla', 'Guatemala', 'Huehuetenango', 'Izabal', 'Jalapa', 'Jutiapa',
                'Petén', 'Quetzaltenango', 'Quiché', 'Retalhuleu', 'Sacatepéquez', 'San Marcos',
                'Santa Rosa', 'Sololá', 'Suchitepéquez', 'Totonicapán', 'Zacapa'
            ],
            'col_inicio': col + 36
        },
        'P44.1 Aduana': {
            'columna': 'P44 - Aduana',
            'categorias': [
                'Central Guatemala', 'El Carmen', 'Integrada Corinto', 'Integrada El Florido',
                'La Mesilla', 'Puerto Barrios Almacenadora Pelícano, S.A -ALPELSA', 'Puerto Quetzal',
                'San Cristóbal', 'Santo Tomás de Castilla Zona Libre de Industria y Comercio -ZOLIC-',
                'Tikal', 'Valle Nuevo'
            ],
            'col_inicio': col + 58
        },
        'P9 Personería': {
            'columna': 'P9 - Personería',
            'categorias': [
                'a. Contribuyente/Propietario.', 'b. Representante Legal', 'c. Abogado y Notario',
                'd. Mandatario', 'e. Contador/auxiliar', 'f. Contador Público y Auditor',
                'g. Gestor Tributario', 'h. Importador', 'i. Exportador', 'j. Asistente de Agente',
                'k. Auxiliar Gestor Tributario', 'm. Consolidador/Descons.', 'n. Transportista Ad',
                'p. Mensajero', 'r. Otro'
            ],
            'col_inicio': col + 69
        },
        'P38 Etnia': {
            'columna': 'P38 - Etnia',
            'categorias': ['Garifuna', 'Ladino', 'Maya', 'Otro', 'Xinca'],
            'col_inicio': col + 84
        },
        'Oficina/Agencia/Delegación': {
            'columna': 'Region_Oficina',
            'categorias': ['Central', 'Occidente', 'Sur', 'Nororiente'],
            'col_inicio': col + 89
        },
        'Aduana': {
            'columna': 'Region_Aduana',
            'categorias': ['Central', 'Occidente', 'Sur', 'Nororiente'],
            'col_inicio': col + 93
        }
    }
    
    # Crear encabezados principales (fila 3)
    col_actual = 1
    ws.cell(row=3, column=col_actual, value='')
    ws.cell(row=3, column=col_actual).fill = fill_header
    ws.cell(row=3, column=col_actual).border = Border(left=medium_side, right=medium_side, top=thin_side, bottom=thin_side)
    col_actual += 1
    
    ws.cell(row=3, column=col_actual, value='TOTAL')
    ws.cell(row=3, column=col_actual).fill = fill_header
    ws.cell(row=3, column=col_actual).border = Border(left=medium_side, right=medium_side, top=thin_side, bottom=thin_side)
    ws.cell(row=3, column=col_actual).font = Font(bold=True)
    ws.cell(row=3, column=col_actual).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    col_actual += 1
    
    for var_nombre, var_info in variables.items():
        num_cols = len(var_info['categorias'])
        inicio = col_actual
        fin = col_actual + num_cols - 1
        
        ws.merge_cells(start_row=3, start_column=inicio, end_row=3, end_column=fin)
        ws.cell(row=3, column=inicio, value=var_nombre)
        ws.cell(row=3, column=inicio).fill = fill_header
        ws.cell(row=3, column=inicio).border = Border(left=medium_side, right=medium_side, top=thin_side, bottom=thin_side)
        ws.cell(row=3, column=inicio).font = Font(bold=True)
        ws.cell(row=3, column=inicio).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        col_actual = fin + 1
    
    # Calcular el número total de columnas (col_actual - 1 es la última columna con datos)
    total_columnas = col_actual - 1
    print(f"Total de columnas calculadas: {total_columnas}")
    
    # Aplicar merge y bordes al título ahora que sabemos el total de columnas
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columnas)
    ws.cell(row=1, column=1).border = Border(
        left=medium_side,
        right=medium_side,
        top=medium_side,
        bottom=thin_side
    )
    
    # Aplicar bordes a la fila vacía (fila 2) solo hasta las columnas necesarias
    for col in range(1, total_columnas + 1):
        cell = ws.cell(row=2, column=col)
        cell.fill = fill_fila1
        cell.border = Border(
            left=thin_side, 
            right=thin_side, 
            top=thin_side, 
            bottom=thin_side
        )
    
    # Fila 4: Sub-encabezados (categorías)
    print("Creando sub-encabezados...")
    col_actual = 1
    ws.cell(row=4, column=col_actual, value='')
    ws.cell(row=4, column=col_actual).border = thin_border
    col_actual += 1
    
    ws.cell(row=4, column=col_actual, value='')
    ws.cell(row=4, column=col_actual).border = Border(
        left=medium_side, 
        right=medium_side, 
        top=Side(style='medium'),  # Cambiado a medium según el ejemplo
        bottom=Side(style='thin', color='FFD0D0D0')
    )
    col_actual += 1
    
    for var_nombre, var_info in variables.items():
        num_cats = len(var_info['categorias'])
        for i, cat in enumerate(var_info['categorias']):
            es_primera = (i == 0)
            es_ultima = (i == num_cats - 1)
            ws.cell(row=4, column=col_actual, value=cat)
            ws.cell(row=4, column=col_actual).font = Font(bold=True)
            border = Border(
                left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),  # Última columna del grupo tiene right=medium
                top=Side(style='medium'),
                bottom=Side(style='thin', color='FFD0D0D0')
            )
            ws.cell(row=4, column=col_actual).border = border
            ws.cell(row=4, column=col_actual).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            col_actual += 1
    
    # Filas de datos: Valores de P4
    print("Generando datos del análisis cruzado...")
    # Solo mostrar las 3 opciones principales, pero incluir todas las combinaciones
    p4_valores = ['a. RTU', 'b. FEL', 'c. Aduanas sin papeles', 'd. Agencia Virtual', 'e. Otros']
    
    print(f"  Opciones de P4 a mostrar: {len(p4_valores)}")
    for opcion in p4_valores:
        # Contar todos los registros con esta opción (P4 no tiene combinaciones)
        count = len(df[df['P4 - Servicio Electrónico'] == opcion])
        print(f"    - {opcion}: {count} registros")
    
    fila = 5
    
    for idx_p4, p4_val in enumerate(p4_valores):
        col_actual = 1
        
        # Nombre de la fila
        ws.cell(row=fila, column=col_actual, value=p4_val)
        border = Border(
            left=Side(style='thin', color='FFD0D0D0'),
            right=Side(style='thin', color='FFD0D0D0'),
            top=Side(style='medium' if idx_p4 == 0 else 'thin', color='FFD0D0D0'),
            bottom=Side(style='thin', color='FFD0D0D0')
        )
        ws.cell(row=fila, column=col_actual).border = border
        ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='left', vertical='center')
        col_actual += 1
        
        # TOTAL - contar todos los registros con esta opción (P4 no tiene combinaciones)
        total = len(df[df['P4 - Servicio Electrónico'] == p4_val])
        ws.cell(row=fila, column=col_actual, value=total)
        border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium' if idx_p4 == 0 else 'thin', color='FFD0D0D0'),
            bottom=Side(style='thin', color='FFD0D0D0')
        )
        ws.cell(row=fila, column=col_actual).border = border
        ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
        col_actual += 1
        
        # Datos por variable
        for var_nombre, var_info in variables.items():
            col_original = var_info['columna']
            num_cats = len(var_info['categorias'])
            
            for i, cat in enumerate(var_info['categorias']):
                es_primera = (i == 0)
                es_ultima = (i == num_cats - 1)
                
                # Contar intersección (P4 no tiene combinaciones, comparación directa)
                if col_original == 'Rango_Edad':
                    count = len(df[(df['P4 - Servicio Electrónico'] == p4_val) & (df[col_original] == cat)])
                elif col_original == 'Region_Oficina' or col_original == 'Region_Aduana':
                    count = len(df[(df['P4 - Servicio Electrónico'] == p4_val) & (df[col_original] == cat)])
                else:
                    count = len(df[(df['P4 - Servicio Electrónico'] == p4_val) & (df[col_original] == cat)])
                valor = count
                
                ws.cell(row=fila, column=col_actual, value=valor)
                border = Border(
                    left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                    right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),  # Última columna del grupo tiene right=medium
                    top=Side(style='medium' if idx_p4 == 0 else 'thin', color='FFD0D0D0'),
                    bottom=Side(style='thin', color='FFD0D0D0')
                )
                ws.cell(row=fila, column=col_actual).border = border
                ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
                col_actual += 1
        
        fila += 1
    
    # Asegurar que la última columna de todas las filas tenga right=medium
    print("Aplicando bordes right=medium a la última columna...")
    for row in range(1, fila + 1):
        cell = ws.cell(row=row, column=total_columnas)
        if cell.border:
            current_border = cell.border
            new_border = Border(
                left=current_border.left,
                right=Side(style='medium'),
                top=current_border.top,
                bottom=current_border.bottom
            )
            cell.border = new_border
        else:
            cell.border = Border(
                left=Side(style='thin', color='FFD0D0D0'),
                right=Side(style='medium'),
                top=Side(style='thin', color='FFD0D0D0'),
                bottom=Side(style='thin', color='FFD0D0D0')
            )
    
    # Fila TOTAL
    print("Generando fila de totales...")
    col_actual = 1
    ws.cell(row=fila, column=col_actual, value='TOTAL')
    ws.cell(row=fila, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila, column=col_actual).border = Border(
        left=Side(style='thin', color='FFD0D0D0'),
        right=Side(style='thin', color='FFD0D0D0'),
        top=Side(style='thin', color='FFD0D0D0'),
        bottom=Side(style='medium')
    )
    ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
    col_actual += 1
    
    # TOTAL general
    total_general = len(df[df['P4 - Servicio Electrónico'].notna()])
    ws.cell(row=fila, column=col_actual, value=total_general)
    ws.cell(row=fila, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila, column=col_actual).border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='thin', color='FFD0D0D0'),
        bottom=Side(style='medium')
    )
    ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
    col_actual += 1
    
    # Totales por categoría (solo para registros con P4)
    df_p4 = df[df['P4 - Servicio Electrónico'].notna()]  # Filtrar solo registros con P4
    
    for var_nombre, var_info in variables.items():
        col_original = var_info['columna']
        num_cats = len(var_info['categorias'])
        
        for i, cat in enumerate(var_info['categorias']):
            es_primera = (i == 0)
            es_ultima = (i == num_cats - 1)
            
            if col_original == 'Rango_Edad':
                total_cat = len(df_p4[df_p4[col_original] == cat])
            elif col_original == 'Region_Oficina' or col_original == 'Region_Aduana':
                total_cat = len(df_p4[df_p4[col_original] == cat])
            else:
                total_cat = len(df_p4[df_p4[col_original] == cat])
            
            ws.cell(row=fila, column=col_actual, value=total_cat)
            ws.cell(row=fila, column=col_actual).font = Font(bold=True)
            border = Border(
                left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),  # Última columna del grupo tiene right=medium
                top=Side(style='thin', color='FFD0D0D0'),
                bottom=Side(style='medium')
            )
            ws.cell(row=fila, column=col_actual).border = border
            ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
            col_actual += 1
    
    # Asegurar que TODAS las celdas de la última fila tengan bottom=medium
    # Solo hasta las columnas necesarias (no más allá)
    print("Aplicando bordes finales a la última fila...")
    for col in range(1, total_columnas + 1):
        cell = ws.cell(row=fila, column=col)
        if cell.border:
            # Mantener los bordes existentes (incluyendo right=medium) pero asegurar bottom=medium
            current_border = cell.border
            # La última columna debe tener right=medium
            right_style = Side(style='medium') if col == total_columnas else current_border.right
            new_border = Border(
                left=current_border.left,
                right=right_style,
                top=current_border.top,
                bottom=Side(style='medium')
            )
            cell.border = new_border
        else:
            # Si no tiene borde, crear uno con bottom=medium
            right_style = Side(style='medium') if col == total_columnas else Side(style='thin', color='FFD0D0D0')
            cell.border = Border(
                left=Side(style='thin', color='FFD0D0D0'),
                right=right_style,
                top=Side(style='thin', color='FFD0D0D0'),
                bottom=Side(style='medium')
            )
    
    # Agregar dos filas vacías
    fila += 2
    
    # ============================================================================
    # TABLA DE PORCENTAJES
    # ============================================================================
    print("Generando tabla de porcentajes...")
    
    # Filtrar solo registros con P4 para cálculos de totales
    df_p4 = df[df['P4 - Servicio Electrónico'].notna()]
    
    # Fila de encabezados principales (igual que la primera tabla)
    fila_porcentajes = fila
    col_actual = 1
    ws.cell(row=fila_porcentajes, column=col_actual, value='')
    ws.cell(row=fila_porcentajes, column=col_actual).fill = fill_header
    ws.cell(row=fila_porcentajes, column=col_actual).border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=thin_side)
    col_actual += 1
    
    ws.cell(row=fila_porcentajes, column=col_actual, value='TOTAL')
    ws.cell(row=fila_porcentajes, column=col_actual).fill = fill_header
    ws.cell(row=fila_porcentajes, column=col_actual).border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=thin_side)
    ws.cell(row=fila_porcentajes, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    col_actual += 1
    
    for var_nombre, var_info in variables.items():
        num_cols = len(var_info['categorias'])
        inicio = col_actual
        fin = col_actual + num_cols - 1
        
        ws.merge_cells(start_row=fila_porcentajes, start_column=inicio, end_row=fila_porcentajes, end_column=fin)
        ws.cell(row=fila_porcentajes, column=inicio, value=var_nombre)
        ws.cell(row=fila_porcentajes, column=inicio).fill = fill_header
        ws.cell(row=fila_porcentajes, column=inicio).border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=thin_side)
        ws.cell(row=fila_porcentajes, column=inicio).font = Font(bold=True)
        ws.cell(row=fila_porcentajes, column=inicio).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        col_actual = fin + 1
    
    # Fila de sub-encabezados
    fila_porcentajes += 1
    col_actual = 1
    ws.cell(row=fila_porcentajes, column=col_actual, value='')
    ws.cell(row=fila_porcentajes, column=col_actual).border = thin_border
    col_actual += 1
    
    ws.cell(row=fila_porcentajes, column=col_actual, value='')
    ws.cell(row=fila_porcentajes, column=col_actual).border = Border(
        left=medium_side, 
        right=medium_side, 
        top=Side(style='medium'), 
        bottom=Side(style='thin', color='FFD0D0D0')
    )
    col_actual += 1
    
    for var_nombre, var_info in variables.items():
        num_cats = len(var_info['categorias'])
        for i, cat in enumerate(var_info['categorias']):
            es_primera = (i == 0)
            es_ultima = (i == num_cats - 1)
            ws.cell(row=fila_porcentajes, column=col_actual, value=cat)
            ws.cell(row=fila_porcentajes, column=col_actual).font = Font(bold=True)
            border = Border(
                left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),
                top=Side(style='medium'),
                bottom=Side(style='thin', color='FFD0D0D0')
            )
            ws.cell(row=fila_porcentajes, column=col_actual).border = border
            ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            col_actual += 1
    
    # Función para redondear porcentajes (0.5 -> 1, 1.5 -> 2, etc.)
    def redondear_porcentaje(valor):
        """Redondea porcentajes: si tiene .5 o más, redondea hacia arriba"""
        if pd.isna(valor) or valor == 0:
            return 0
        # Si el decimal es >= 0.5, redondear hacia arriba
        if valor % 1 >= 0.5:
            return int(valor) + 1
        else:
            return int(valor)
    
    # Filas de datos con porcentajes
    fila_porcentajes += 1
    for idx_p4, p4_val in enumerate(p4_valores):
        col_actual = 1
        
        # Nombre de la fila
        ws.cell(row=fila_porcentajes, column=col_actual, value=p4_val)
        border = Border(
            left=Side(style='thin', color='FFD0D0D0'),
            right=Side(style='thin', color='FFD0D0D0'),
            top=Side(style='medium' if idx_p4 == 0 else 'thin', color='FFD0D0D0'),
            bottom=Side(style='thin', color='FFD0D0D0')
        )
        ws.cell(row=fila_porcentajes, column=col_actual).border = border
        ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='left', vertical='center')
        col_actual += 1
        
        # TOTAL - calcular porcentaje sobre el total general
        total_absoluto = len(df[df['P4 - Servicio Electrónico'] == p4_val])
        total_general = len(df[df['P4 - Servicio Electrónico'].notna()])
        porcentaje_total = (total_absoluto / total_general * 100) if total_general > 0 else 0
        porcentaje_redondeado = redondear_porcentaje(porcentaje_total)
        if porcentaje_redondeado == 0:
            ws.cell(row=fila_porcentajes, column=col_actual, value="---")
        else:
            # Guardar como número decimal para formato de porcentaje
            ws.cell(row=fila_porcentajes, column=col_actual, value=porcentaje_redondeado / 100)
            ws.cell(row=fila_porcentajes, column=col_actual).number_format = '0%'
        border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium' if idx_p4 == 0 else 'thin', color='FFD0D0D0'),
            bottom=Side(style='thin', color='FFD0D0D0')
        )
        ws.cell(row=fila_porcentajes, column=col_actual).border = border
        ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
        col_actual += 1
        
        # Datos por variable (porcentajes)
        for var_nombre, var_info in variables.items():
            col_original = var_info['columna']
            num_cats = len(var_info['categorias'])
            
            for i, cat in enumerate(var_info['categorias']):
                es_primera = (i == 0)
                es_ultima = (i == num_cats - 1)
                
                # Contar intersección (P4 no tiene combinaciones)
                if col_original == 'Rango_Edad':
                    count = len(df[(df['P4 - Servicio Electrónico'] == p4_val) & (df[col_original] == cat)])
                elif col_original == 'Region_Oficina' or col_original == 'Region_Aduana':
                    count = len(df[(df['P4 - Servicio Electrónico'] == p4_val) & (df[col_original] == cat)])
                else:
                    count = len(df[(df['P4 - Servicio Electrónico'] == p4_val) & (df[col_original] == cat)])
                
                # Calcular porcentaje VERTICAL (sobre el total de esa categoría/columna - solo registros con P4)
                if col_original == 'Rango_Edad':
                    total_categoria = len(df_p4[df_p4[col_original] == cat])
                elif col_original == 'Region_Oficina' or col_original == 'Region_Aduana':
                    total_categoria = len(df_p4[df_p4[col_original] == cat])
                else:
                    total_categoria = len(df_p4[df_p4[col_original] == cat])
                
                porcentaje = (count / total_categoria * 100) if total_categoria > 0 else 0
                
                porcentaje_redondeado = redondear_porcentaje(porcentaje)
                if porcentaje_redondeado == 0:
                    ws.cell(row=fila_porcentajes, column=col_actual, value="---")
                else:
                    # Guardar como número decimal para formato de porcentaje
                    ws.cell(row=fila_porcentajes, column=col_actual, value=porcentaje_redondeado / 100)
                    ws.cell(row=fila_porcentajes, column=col_actual).number_format = '0%'
                border = Border(
                    left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                    right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),
                    top=Side(style='medium' if idx_p4 == 0 else 'thin', color='FFD0D0D0'),
                    bottom=Side(style='thin', color='FFD0D0D0')
                )
                ws.cell(row=fila_porcentajes, column=col_actual).border = border
                ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
                col_actual += 1
        
        fila_porcentajes += 1
    
    # Fila TOTAL de porcentajes
    col_actual = 1
    ws.cell(row=fila_porcentajes, column=col_actual, value='TOTAL')
    ws.cell(row=fila_porcentajes, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila_porcentajes, column=col_actual).border = Border(
        left=Side(style='thin', color='FFD0D0D0'),
        right=Side(style='thin', color='FFD0D0D0'),
        top=Side(style='thin', color='FFD0D0D0'),
        bottom=Side(style='medium')
    )
    ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
    col_actual += 1
    
    # TOTAL general - suma vertical de porcentajes
    # Sumar los porcentajes de las 5 filas anteriores (opciones de P4)
    suma_total = 0
    for idx_p4 in range(len(p4_valores)):
        fila_anterior = fila_porcentajes - len(p4_valores) + idx_p4
        valor_celda = ws.cell(row=fila_anterior, column=col_actual).value
        if valor_celda is not None:
            if isinstance(valor_celda, (int, float)):
                # Es un número decimal (formato de porcentaje), convertir a porcentaje entero
                suma_total += int(valor_celda * 100)
            elif isinstance(valor_celda, str) and valor_celda != "---":
                # Formato antiguo con "%" (por compatibilidad)
                if '%' in valor_celda:
                    num = int(valor_celda.replace('%', ''))
                    suma_total += num
            # Si es "---", se trata como 0 (no se suma nada)
    
    if suma_total == 0:
        ws.cell(row=fila_porcentajes, column=col_actual, value="---")
    else:
        # Guardar como número decimal para formato de porcentaje
        ws.cell(row=fila_porcentajes, column=col_actual, value=suma_total / 100)
        ws.cell(row=fila_porcentajes, column=col_actual).number_format = '0%'
    ws.cell(row=fila_porcentajes, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila_porcentajes, column=col_actual).border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='thin', color='FFD0D0D0'),
        bottom=Side(style='medium')
    )
    ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
    col_actual += 1
    
    # Totales por categoría - suma vertical de porcentajes
    for var_nombre, var_info in variables.items():
        col_original = var_info['columna']
        num_cats = len(var_info['categorias'])
        
        for i, cat in enumerate(var_info['categorias']):
            es_primera = (i == 0)
            es_ultima = (i == num_cats - 1)
            
            # Sumar los porcentajes verticalmente de las 5 filas anteriores
            suma_porcentajes = 0
            for idx_p4 in range(len(p4_valores)):
                fila_anterior = fila_porcentajes - len(p4_valores) + idx_p4
                valor_celda = ws.cell(row=fila_anterior, column=col_actual).value
                if valor_celda is not None:
                    if isinstance(valor_celda, (int, float)):
                        # Es un número decimal (formato de porcentaje), convertir a porcentaje entero
                        suma_porcentajes += int(valor_celda * 100)
                    elif isinstance(valor_celda, str) and valor_celda != "---":
                        # Formato antiguo con "%" (por compatibilidad)
                        if '%' in valor_celda:
                            num = int(valor_celda.replace('%', ''))
                            suma_porcentajes += num
                    # Si es "---", se trata como 0 (no se suma nada)
            
            if suma_porcentajes == 0:
                ws.cell(row=fila_porcentajes, column=col_actual, value="---")
            else:
                # Guardar como número decimal para formato de porcentaje
                ws.cell(row=fila_porcentajes, column=col_actual, value=suma_porcentajes / 100)
                ws.cell(row=fila_porcentajes, column=col_actual).number_format = '0%'
            ws.cell(row=fila_porcentajes, column=col_actual).font = Font(bold=True)
            border = Border(
                left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),
                top=Side(style='thin', color='FFD0D0D0'),
                bottom=Side(style='medium')
            )
            ws.cell(row=fila_porcentajes, column=col_actual).border = border
            ws.cell(row=fila_porcentajes, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
            col_actual += 1
    
    # Asegurar bordes en la última fila y última columna de porcentajes
    for col in range(1, total_columnas + 1):
        cell = ws.cell(row=fila_porcentajes, column=col)
        if cell.border:
            current_border = cell.border
            right_style = Side(style='medium') if col == total_columnas else current_border.right
            new_border = Border(
                left=current_border.left,
                right=right_style,
                top=current_border.top,
                bottom=Side(style='medium')
            )
            cell.border = new_border
    
    # Asegurar que la última columna de todas las filas de porcentajes tenga right=medium
    for row in range(fila - 2, fila_porcentajes + 1):
        cell = ws.cell(row=row, column=total_columnas)
        if cell.border:
            current_border = cell.border
            new_border = Border(
                left=current_border.left,
                right=Side(style='medium'),
                top=current_border.top,
                bottom=current_border.bottom
            )
            cell.border = new_border
    
    # Ajustar ancho de columnas solo hasta las necesarias
    print("Ajustando ancho de columnas...")
    ws.column_dimensions['A'].width = 30
    for col in range(2, total_columnas + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    # Guardar archivo
    print(f"Guardando archivo: {archivo_salida}")
    try:
        wb.save(archivo_salida)
        print(f"✓ Archivo generado exitosamente: {archivo_salida}")
        print(f"  Total de registros procesados: {len(df)}")
        print(f"  Total de filas en el análisis: {fila}")
        print(f"  Total de columnas: {ws.max_column}")
    except Exception as e:
        print(f"ERROR al guardar el archivo: {e}")
        sys.exit(1)

if __name__ == "__main__":
    # Permitir especificar archivos como argumentos
    archivo_entrada = sys.argv[1] if len(sys.argv) > 1 else 'V3.xlsx'
    archivo_salida = sys.argv[2] if len(sys.argv) > 2 else 'P4-Cruzado.xlsx'
    
    print("=" * 60)
    print("GENERADOR DE ANÁLISIS CRUZADO P4")
    print("=" * 60)
    print()
    
    generar_analisis_cruzado(archivo_entrada, archivo_salida)
    
    print()
    print("=" * 60)
    print("Proceso completado exitosamente")
    print("=" * 60)

