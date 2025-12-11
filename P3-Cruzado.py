#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para generar análisis cruzado de la pregunta P3 (Medios SAT Utilizados)
con todas las variables demográficas y de clasificación.

Autor: Generado automáticamente
Fecha: 2025
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys

def normalizar_p3(valor):
    """
    Normaliza los valores de P3 manteniendo todas las combinaciones.
    Normaliza el orden para que combinaciones equivalentes se agrupen correctamente.
    """
    if pd.isna(valor):
        return None
    valor_str = str(valor).strip()
    
    # Detectar qué opciones están presentes
    tiene_presencial = 'a. Presencial' in valor_str
    tiene_contact = 'b. Contact Center' in valor_str
    tiene_electronicos = 'c. Servicios Electrónicos' in valor_str
    
    # Construir la combinación normalizada en orden estándar
    opciones = []
    if tiene_presencial:
        opciones.append('a. Presencial')
    if tiene_contact:
        opciones.append('b. Contact Center')
    if tiene_electronicos:
        opciones.append('c. Servicios Electrónicos')
    
    if len(opciones) == 0:
        return None
    
    # Retornar combinación normalizada (siempre en el mismo orden)
    return ', '.join(opciones)

def crear_rango_edad(edad):
    """
    Crea rangos de edad a partir de la edad numérica.
    """
    if pd.isna(edad):
        return None
    try:
        edad_num = int(float(edad))
        if edad_num <= 25:
            return '18 - 25'
        elif edad_num <= 35:
            return '26 - 35'
        elif edad_num <= 45:
            return '36 - 45'
        elif edad_num <= 60:
            return '46 - 60'
        else:
            return 'Más de 61'
    except:
        return None

def obtener_region_oficina(oficina):
    """
    Agrupa las oficinas/agencias/delegaciones por región.
    """
    if pd.isna(oficina):
        return None
    
    oficina_str = str(oficina).strip()
    
    # Central
    central = ['Chimaltenango', 'El Progreso', 'Guatemala', 'Sacatepéquez']
    if oficina_str in central:
        return 'Central'
    
    # Occidente
    occidente = ['Huehuetenango', 'Quetzaltenango', 'Quiché', 'San Marcos', 'Sololá', 'Totonicapán']
    if oficina_str in occidente:
        return 'Occidente'
    
    # Sur
    sur = ['Escuintla', 'Jutiapa', 'Retalhuleu', 'Santa Rosa', 'Suchitepéquez']
    if oficina_str in sur:
        return 'Sur'
    
    # Nororiente
    nororiente = ['Alta Verapaz', 'Baja Verapaz', 'Chiquimula', 'Izabal', 'Jalapa', 'Petén', 'Zacapa']
    if oficina_str in nororiente:
        return 'Nororiente'
    
    return None

def obtener_region_aduana(aduana):
    """
    Agrupa las aduanas por región.
    """
    if pd.isna(aduana):
        return None
    
    aduana_str = str(aduana).strip()
    
    # Central
    if 'Central Guatemala' in aduana_str:
        return 'Central'
    
    # Occidente
    if 'El Carmen' in aduana_str or 'La Mesilla' in aduana_str:
        return 'Occidente'
    
    # Sur
    if 'San Cristóbal' in aduana_str or 'Valle Nuevo' in aduana_str or 'Puerto Quetzal' in aduana_str:
        return 'Sur'
    
    # Nororiente
    if 'Integrada Corinto' in aduana_str or 'Integrada El Florido' in aduana_str or 'Puerto Barrios' in aduana_str or 'Santo Tomás' in aduana_str or 'Tikal' in aduana_str:
        return 'Nororiente'
    
    return None

def aplicar_estilos_bordes(ws, fila, col, es_primera_fila=False, es_ultima_fila=False, 
                           es_primera_col=False, es_ultima_col=False, es_encabezado=False):
    """
    Aplica los estilos de bordes según la posición de la celda.
    """
    thin_side = Side(style='thin', color='FFD0D0D0')
    medium_side = Side(style='medium')
    
    left_style = medium_side if es_primera_col else thin_side
    right_style = medium_side if es_ultima_col else thin_side
    top_style = medium_side if es_primera_fila else thin_side
    bottom_style = medium_side if es_ultima_fila else thin_side
    
    border = Border(
        left=left_style,
        right=right_style,
        top=top_style,
        bottom=bottom_style
    )
    
    ws.cell(row=fila, column=col).border = border
    return border

def generar_analisis_cruzado(archivo_entrada='V3.xlsx', archivo_salida='Analisis_Cruzado_P3.xlsx'):
    """
    Función principal que genera el análisis cruzado de P3.
    
    Args:
        archivo_entrada: Nombre del archivo Excel de entrada (default: V3.xlsx)
        archivo_salida: Nombre del archivo Excel de salida (default: Analisis_Cruzado_P3.xlsx)
    """
    print(f"Leyendo archivo: {archivo_entrada}")
    
    # Verificar que el archivo existe
    if not os.path.exists(archivo_entrada):
        print(f"ERROR: No se encontró el archivo {archivo_entrada}")
        sys.exit(1)
    
    # Leer el archivo
    try:
        df = pd.read_excel(archivo_entrada)
        print(f"Archivo leído exitosamente. Total de registros: {len(df)}")
    except Exception as e:
        print(f"ERROR al leer el archivo: {e}")
        sys.exit(1)
    
    # Crear columna normalizada de P3
    print("Normalizando valores de P3...")
    df['P3_norm'] = df['P3 - Medios SAT Utilizados'].apply(normalizar_p3)
    
    # Crear rangos de edad
    print("Creando rangos de edad...")
    df['Rango_Edad'] = df['P36 - Edad'].apply(crear_rango_edad)
    
    # Crear regiones para Oficina/Agencia/Delegación
    print("Creando regiones de Oficina/Agencia/Delegación...")
    df['Region_Oficina'] = df['P44 - Oficina/Agencia/Delegación'].apply(obtener_region_oficina)
    
    # Crear regiones para Aduana
    print("Creando regiones de Aduana...")
    df['Region_Aduana'] = df['P44 - Aduana'].apply(obtener_region_aduana)
    
    # Crear nuevo workbook
    print("Creando estructura del archivo Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "P3"
    
    # Definir estilos
    thin_side = Side(style='thin', color='FFD0D0D0')
    medium_side = Side(style='medium')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    fill_fila1 = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    fill_header = PatternFill(start_color='FFE7E6E6', end_color='FFE7E6E6', fill_type='solid')
    
    # Fila 0: Título de la pregunta (se ajustará después de calcular columnas)
    print("Agregando título de la pregunta...")
    ws.cell(row=1, column=1, value='P3 - Medios SAT Utilizados')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')  # Alineado a la izquierda
    ws.cell(row=1, column=1).fill = fill_fila1
    
    # Fila 1: Vacía con fondo gris (se ajustará después de calcular total_columnas)
    print("Configurando formato de la primera fila...")
    
    # Fila 3: Encabezados principales
    print("Creando encabezados principales...")
    col = 1
    ws.cell(row=3, column=col, value='')
    ws.cell(row=3, column=col).fill = fill_header
    ws.cell(row=3, column=col).border = Border(left=medium_side, right=medium_side, top=thin_side, bottom=thin_side)
    col += 1
    
    # TOTAL
    ws.cell(row=3, column=col, value='TOTAL')
    ws.cell(row=3, column=col).fill = fill_header
    ws.cell(row=3, column=col).border = Border(left=medium_side, right=medium_side, top=thin_side, bottom=thin_side)
    ws.cell(row=3, column=col).font = Font(bold=True)
    ws.cell(row=3, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    col += 1
    
    # Definir todas las variables y sus categorías
    variables = {
        'P37 Género': {
            'columna': 'P37 - Género',
            'categorias': ['H', 'M', 'No deseo responder'],
            'col_inicio': col
        },
        'Rango de edad': {
            'columna': 'Rango_Edad',
            'categorias': ['18 - 25', '26 - 35', '36 - 45', '46 - 60', 'Más de 61'],
            'col_inicio': col + 3
        },
        'P40 Nivel académico': {
            'columna': 'P40 - Nivel Académico',
            'categorias': [
                'a. Ninguno', 'b. Primaria incompleta', 'c. Primaria completa',
                'd. Secundaria incompleta (1ro a 3ro básico)', 'e. Secundaria Completa (1ro a 3ro básico)',
                'f. Diversificado incompleto', 'g. Diversificado completo', 'h. Técnico',
                'i. Universidad incompleta', 'j. Universidad Completa', 'k. Maestría / Posgrado'
            ],
            'col_inicio': col + 8
        },
        'P39 Idiomas': {
            'columna': 'P39 - Idiomas',
            'categorias': [
                'a. Achi', 'b. Qánjob\'al', 'c. Q\'eqchi', 'd. Akateco', 'e. Kaqchikel',
                'f. Sakapulteko', 'h. Kiché', 'i. Sipakapense', 'k. Mam', 'n. Mopan',
                'p. Ixil', 'q. Poqomam', 's. Jakalteco', 't. Poqomchi', 'u. Ninguno',
                'v. Inglés', 'w. Otro'
            ],
            'col_inicio': col + 19
        },
        'P44 Oficina/Agencia/Delegación': {
            'columna': 'P44 - Oficina/Agencia/Delegación',
            'categorias': [
                'Alta Verapaz', 'Baja Verapaz', 'Chimaltenango', 'Chiquimula', 'El Progreso',
                'Escuintla', 'Guatemala', 'Huehuetenango', 'Izabal', 'Jalapa', 'Jutiapa',
                'Petén', 'Quetzaltenango', 'Quiché', 'Retalhuleu', 'Sacatepéquez', 'San Marcos',
                'Santa Rosa', 'Sololá', 'Suchitepéquez', 'Totonicapán', 'Zacapa'
            ],
            'col_inicio': col + 36
        },
        'P44.1 Aduana': {
            'columna': 'P44 - Aduana',
            'categorias': [
                'Central Guatemala', 'El Carmen', 'Integrada Corinto', 'Integrada El Florido',
                'La Mesilla', 'Puerto Barrios Almacenadora Pelícano, S.A -ALPELSA', 'Puerto Quetzal',
                'San Cristóbal', 'Santo Tomás de Castilla Zona Libre de Industria y Comercio -ZOLIC-',
                'Tikal', 'Valle Nuevo'
            ],
            'col_inicio': col + 58
        },
        'P9 Personería': {
            'columna': 'P9 - Personería',
            'categorias': [
                'a. Contribuyente/Propietario.', 'b. Representante Legal', 'c. Abogado y Notario',
                'd. Mandatario', 'e. Contador/auxiliar', 'f. Contador Público y Auditor',
                'g. Gestor Tributario', 'h. Importador', 'i. Exportador', 'j. Asistente de Agente',
                'k. Auxiliar Gestor Tributario', 'm. Consolidador/Descons.', 'n. Transportista Ad',
                'p. Mensajero', 'r. Otro'
            ],
            'col_inicio': col + 69
        },
        'P38 Etnia': {
            'columna': 'P38 - Etnia',
            'categorias': ['Garifuna', 'Ladino', 'Maya', 'Otro', 'Xinca'],
            'col_inicio': col + 84
        },
        'P3 Medios SAT utilizados': {
            'columna': 'P3_norm',
            'categorias': ['a. Presencial', 'b. Contact Center', 'c. Servicios Electrónicos'],
            'col_inicio': col + 89
        },
        'Oficina/Agencia/Delegación': {
            'columna': 'Region_Oficina',
            'categorias': ['Central', 'Occidente', 'Sur', 'Nororiente'],
            'col_inicio': col + 92
        },
        'Aduana': {
            'columna': 'Region_Aduana',
            'categorias': ['Central', 'Occidente', 'Sur', 'Nororiente'],
            'col_inicio': col + 96
        }
    }
    
    # Crear encabezados principales (fila 3)
    col_actual = 1
    ws.cell(row=3, column=col_actual, value='')
    ws.cell(row=3, column=col_actual).fill = fill_header
    ws.cell(row=3, column=col_actual).border = Border(left=medium_side, right=medium_side, top=thin_side, bottom=thin_side)
    col_actual += 1
    
    ws.cell(row=3, column=col_actual, value='TOTAL')
    ws.cell(row=3, column=col_actual).fill = fill_header
    ws.cell(row=3, column=col_actual).border = Border(left=medium_side, right=medium_side, top=thin_side, bottom=thin_side)
    ws.cell(row=3, column=col_actual).font = Font(bold=True)
    ws.cell(row=3, column=col_actual).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    col_actual += 1
    
    for var_nombre, var_info in variables.items():
        num_cols = len(var_info['categorias'])
        inicio = col_actual
        fin = col_actual + num_cols - 1
        
        ws.merge_cells(start_row=3, start_column=inicio, end_row=3, end_column=fin)
        ws.cell(row=3, column=inicio, value=var_nombre)
        ws.cell(row=3, column=inicio).fill = fill_header
        ws.cell(row=3, column=inicio).border = Border(left=medium_side, right=medium_side, top=thin_side, bottom=thin_side)
        ws.cell(row=3, column=inicio).font = Font(bold=True)
        ws.cell(row=3, column=inicio).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        col_actual = fin + 1
    
    # Calcular el número total de columnas (col_actual - 1 es la última columna con datos)
    total_columnas = col_actual - 1
    print(f"Total de columnas calculadas: {total_columnas}")
    
    # Aplicar merge y bordes al título ahora que sabemos el total de columnas
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columnas)
    ws.cell(row=1, column=1).border = Border(
        left=medium_side,
        right=medium_side,
        top=medium_side,
        bottom=thin_side
    )
    
    # Aplicar bordes a la fila vacía (fila 2) solo hasta las columnas necesarias
    for col in range(1, total_columnas + 1):
        cell = ws.cell(row=2, column=col)
        cell.fill = fill_fila1
        cell.border = Border(
            left=thin_side, 
            right=thin_side, 
            top=thin_side, 
            bottom=thin_side
        )
    
    # Fila 4: Sub-encabezados (categorías)
    print("Creando sub-encabezados...")
    col_actual = 1
    ws.cell(row=4, column=col_actual, value='')
    ws.cell(row=4, column=col_actual).border = thin_border
    col_actual += 1
    
    ws.cell(row=4, column=col_actual, value='')
    ws.cell(row=4, column=col_actual).border = Border(
        left=medium_side, 
        right=medium_side, 
        top=Side(style='medium'),  # Cambiado a medium según el ejemplo
        bottom=Side(style='thin', color='FFD0D0D0')
    )
    col_actual += 1
    
    for var_nombre, var_info in variables.items():
        num_cats = len(var_info['categorias'])
        for i, cat in enumerate(var_info['categorias']):
            es_primera = (i == 0)
            es_ultima = (i == num_cats - 1)
            ws.cell(row=4, column=col_actual, value=cat)
            ws.cell(row=4, column=col_actual).font = Font(bold=True)
            border = Border(
                left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),  # Última columna del grupo tiene right=medium
                top=Side(style='medium'),
                bottom=Side(style='thin', color='FFD0D0D0')
            )
            ws.cell(row=4, column=col_actual).border = border
            ws.cell(row=4, column=col_actual).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            col_actual += 1
    
    # Filas de datos: Valores de P3
    print("Generando datos del análisis cruzado...")
    # Solo mostrar las 3 opciones principales, pero incluir todas las combinaciones
    p3_valores = ['a. Presencial', 'b. Contact Center', 'c. Servicios Electrónicos']
    
    print(f"  Opciones de P3 a mostrar: {len(p3_valores)}")
    for opcion in p3_valores:
        # Contar todos los registros que contengan esta opción (incluyendo combinaciones)
        count = len(df[df['P3_norm'].str.contains(opcion, na=False)])
        print(f"    - {opcion}: {count} registros (incluyendo combinaciones)")
    
    fila = 5
    
    for idx_p3, p3_val in enumerate(p3_valores):
        col_actual = 1
        
        # Nombre de la fila
        ws.cell(row=fila, column=col_actual, value=p3_val)
        border = Border(
            left=Side(style='thin', color='FFD0D0D0'),
            right=Side(style='thin', color='FFD0D0D0'),
            top=Side(style='medium' if idx_p3 == 0 else 'thin', color='FFD0D0D0'),
            bottom=Side(style='thin', color='FFD0D0D0')
        )
        ws.cell(row=fila, column=col_actual).border = border
        ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='left', vertical='center')
        col_actual += 1
        
        # TOTAL - contar todos los registros que contengan esta opción
        # Usar la columna original sin normalizar para que coincida con el ejemplo
        total = len(df[df['P3 - Medios SAT Utilizados'].str.contains(p3_val, na=False)])
        ws.cell(row=fila, column=col_actual, value=total)
        border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium' if idx_p3 == 0 else 'thin', color='FFD0D0D0'),
            bottom=Side(style='thin', color='FFD0D0D0')
        )
        ws.cell(row=fila, column=col_actual).border = border
        ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
        col_actual += 1
        
        # Datos por variable
        for var_nombre, var_info in variables.items():
            col_original = var_info['columna']
            num_cats = len(var_info['categorias'])
            
            for i, cat in enumerate(var_info['categorias']):
                es_primera = (i == 0)
                es_ultima = (i == num_cats - 1)
                
                if var_nombre == 'P3 Medios SAT utilizados':
                    # Para la columna de P3, verificar si p3_val contiene esta opción
                    # p3_val es una opción individual como "a. Presencial"
                    # cat es también una opción individual como "a. Presencial"
                    if p3_val == cat:
                        valor = total
                    else:
                        valor = 0
                else:
                    # Contar intersección - incluir todas las combinaciones que contengan p3_val
                    if col_original == 'Rango_Edad':
                        count = len(df[(df['P3_norm'].str.contains(p3_val, na=False)) & (df[col_original] == cat)])
                    else:
                        count = len(df[(df['P3_norm'].str.contains(p3_val, na=False)) & (df[col_original] == cat)])
                    valor = count
                
                ws.cell(row=fila, column=col_actual, value=valor)
                border = Border(
                    left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                    right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),  # Última columna del grupo tiene right=medium
                    top=Side(style='medium' if idx_p3 == 0 else 'thin', color='FFD0D0D0'),
                    bottom=Side(style='thin', color='FFD0D0D0')
                )
                ws.cell(row=fila, column=col_actual).border = border
                ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
                col_actual += 1
        
        fila += 1
    
    # Asegurar que la última columna de todas las filas tenga right=medium
    print("Aplicando bordes right=medium a la última columna...")
    for row in range(1, fila + 1):
        cell = ws.cell(row=row, column=total_columnas)
        if cell.border:
            current_border = cell.border
            new_border = Border(
                left=current_border.left,
                right=Side(style='medium'),
                top=current_border.top,
                bottom=current_border.bottom
            )
            cell.border = new_border
        else:
            cell.border = Border(
                left=Side(style='thin', color='FFD0D0D0'),
                right=Side(style='medium'),
                top=Side(style='thin', color='FFD0D0D0'),
                bottom=Side(style='thin', color='FFD0D0D0')
            )
    
    # Fila TOTAL
    print("Generando fila de totales...")
    col_actual = 1
    ws.cell(row=fila, column=col_actual, value='TOTAL')
    ws.cell(row=fila, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila, column=col_actual).border = Border(
        left=Side(style='thin', color='FFD0D0D0'),
        right=Side(style='thin', color='FFD0D0D0'),
        top=Side(style='thin', color='FFD0D0D0'),
        bottom=Side(style='medium')
    )
    ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
    col_actual += 1
    
    # TOTAL general
    total_general = len(df[df['P3_norm'].notna()])
    ws.cell(row=fila, column=col_actual, value=total_general)
    ws.cell(row=fila, column=col_actual).font = Font(bold=True)
    ws.cell(row=fila, column=col_actual).border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='thin', color='FFD0D0D0'),
        bottom=Side(style='medium')
    )
    ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
    col_actual += 1
    
    # Totales por categoría
    for var_nombre, var_info in variables.items():
        col_original = var_info['columna']
        num_cats = len(var_info['categorias'])
        
        for i, cat in enumerate(var_info['categorias']):
            es_primera = (i == 0)
            es_ultima = (i == num_cats - 1)
            
            if var_nombre == 'P3 Medios SAT utilizados':
                # Contar todos los registros que contengan esta opción (incluyendo combinaciones)
                total_cat = len(df[df[col_original].str.contains(cat, na=False)])
            else:
                if col_original == 'Rango_Edad':
                    total_cat = len(df[df[col_original] == cat])
                else:
                    total_cat = len(df[df[col_original] == cat])
            
            ws.cell(row=fila, column=col_actual, value=total_cat)
            ws.cell(row=fila, column=col_actual).font = Font(bold=True)
            border = Border(
                left=Side(style='medium' if es_primera else 'thin', color='FFD0D0D0'),
                right=Side(style='medium' if es_ultima else 'thin', color='FFD0D0D0'),  # Última columna del grupo tiene right=medium
                top=Side(style='thin', color='FFD0D0D0'),
                bottom=Side(style='medium')
            )
            ws.cell(row=fila, column=col_actual).border = border
            ws.cell(row=fila, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
            col_actual += 1
    
    # Asegurar que TODAS las celdas de la última fila tengan bottom=medium
    # Solo hasta las columnas necesarias (no más allá)
    print("Aplicando bordes finales a la última fila...")
    for col in range(1, total_columnas + 1):
        cell = ws.cell(row=fila, column=col)
        if cell.border:
            # Mantener los bordes existentes (incluyendo right=medium) pero asegurar bottom=medium
            current_border = cell.border
            # La última columna debe tener right=medium
            right_style = Side(style='medium') if col == total_columnas else current_border.right
            new_border = Border(
                left=current_border.left,
                right=right_style,
                top=current_border.top,
                bottom=Side(style='medium')
            )
            cell.border = new_border
        else:
            # Si no tiene borde, crear uno con bottom=medium
            right_style = Side(style='medium') if col == total_columnas else Side(style='thin', color='FFD0D0D0')
            cell.border = Border(
                left=Side(style='thin', color='FFD0D0D0'),
                right=right_style,
                top=Side(style='thin', color='FFD0D0D0'),
                bottom=Side(style='medium')
            )
    
    # Ajustar ancho de columnas solo hasta las necesarias
    print("Ajustando ancho de columnas...")
    ws.column_dimensions['A'].width = 30
    for col in range(2, total_columnas + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    # Guardar archivo
    print(f"Guardando archivo: {archivo_salida}")
    try:
        wb.save(archivo_salida)
        print(f"✓ Archivo generado exitosamente: {archivo_salida}")
        print(f"  Total de registros procesados: {len(df)}")
        print(f"  Total de filas en el análisis: {fila}")
        print(f"  Total de columnas: {ws.max_column}")
    except Exception as e:
        print(f"ERROR al guardar el archivo: {e}")
        sys.exit(1)

if __name__ == "__main__":
    # Permitir especificar archivos como argumentos
    archivo_entrada = sys.argv[1] if len(sys.argv) > 1 else 'V3.xlsx'
    archivo_salida = sys.argv[2] if len(sys.argv) > 2 else 'P3-Cruzado.xlsx'
    
    print("=" * 60)
    print("GENERADOR DE ANÁLISIS CRUZADO P3")
    print("=" * 60)
    print()
    
    generar_analisis_cruzado(archivo_entrada, archivo_salida)
    
    print()
    print("=" * 60)
    print("Proceso completado exitosamente")
    print("=" * 60)

